from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError:  # pragma: no cover - exercised by user environment
    load_workbook = None
    get_column_letter = None


SUPPORTED_OPENPYXL_EXTENSIONS = {".xlsx", ".xlsm"}
KNOWN_WORKBOOK_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
DEFAULT_HEADER_SCAN_ROWS = 50
DEFAULT_SAMPLE_ROWS = 3
DEFAULT_SAMPLE_COLUMNS = 12
MAX_CELL_TEXT = 80


HEADER_PATTERNS = {
    "distributor_sku": [
        "distributor sku",
        "vendor sku",
        "vendor part",
        "vendor part number",
        "item number",
        "item no",
        "item #",
        "sku",
        "catalog number",
        "catalog no",
        "product code",
    ],
    "manufacturer_part_number": [
        "manufacturer part number",
        "manufacturer part",
        "manufacturer item",
        "mfg part",
        "mfr part",
        "mpn",
        "model number",
        "model no",
    ],
    "description": [
        "description",
        "product description",
        "item description",
        "item name",
        "product name",
        "title",
    ],
    "uom": [
        "uom",
        "unit of measure",
        "unit",
        "units",
        "sell unit",
        "selling unit",
    ],
    "price": [
        "price",
        "contract price",
        "net price",
        "dealer price",
        "distributor price",
        "cost",
        "list price",
        "msrp",
        "your price",
    ],
    "effective_date": [
        "effective date",
        "start date",
        "valid from",
        "price effective",
        "eff date",
    ],
    "expiration_date": [
        "expiration date",
        "expiry date",
        "end date",
        "valid through",
        "valid until",
        "expires",
    ],
}


@dataclass
class ColumnCandidate:
    column_index: int
    column_letter: str
    header: str
    confidence: float
    matched_term: str


@dataclass
class CandidateHeaderRow:
    sheet_name: str
    row_index: int
    score: float
    non_empty_count: int
    matched_categories: list[str]
    labels: list[str]


@dataclass
class SheetProfile:
    name: str
    state: str
    dimensions: dict[str, int]
    used_range: dict[str, int | str | None]
    merged_cell_ranges: list[str]
    formula_count: int
    hidden_row_count: int
    hidden_column_count: int
    candidate_header_rows: list[CandidateHeaderRow] = field(default_factory=list)
    candidate_pricing_sheet_score: float = 0
    candidate_pricing_sheet_reasons: list[str] = field(default_factory=list)
    detected_header_labels: list[str] = field(default_factory=list)
    likely_distributor_sku_columns: list[ColumnCandidate] = field(default_factory=list)
    likely_manufacturer_part_number_columns: list[ColumnCandidate] = field(default_factory=list)
    likely_description_columns: list[ColumnCandidate] = field(default_factory=list)
    likely_uom_columns: list[ColumnCandidate] = field(default_factory=list)
    likely_price_columns: list[ColumnCandidate] = field(default_factory=list)
    likely_effective_date_columns: list[ColumnCandidate] = field(default_factory=list)
    likely_expiration_date_columns: list[ColumnCandidate] = field(default_factory=list)
    sample_rows: list[dict[str, Any]] = field(default_factory=list)
    risk_flags: list[str] = field(default_factory=list)


@dataclass
class WorkbookProfile:
    file_name: str
    file_hash: str
    file_size: int
    workbook_extension: str
    manifest_metadata: dict[str, str] | None
    sheet_names: list[str] = field(default_factory=list)
    hidden_sheets: list[str] = field(default_factory=list)
    sheets: list[SheetProfile] = field(default_factory=list)
    candidate_pricing_sheets: list[dict[str, Any]] = field(default_factory=list)
    risk_flags: list[str] = field(default_factory=list)


def normalize_label(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"[_\-/#:]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip().lower()


def display_cell_value(value: Any) -> str | int | float | bool | None:
    if value is None:
        return None
    if isinstance(value, (int, float, bool)):
        return value
    text = str(value).strip()
    if len(text) > MAX_CELL_TEXT:
        return f"{text[:MAX_CELL_TEXT]}..."
    return text


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_manifest(path: Path | None) -> dict[str, dict[str, str]]:
    if path is None or not path.exists():
        return {}

    manifest: dict[str, dict[str, str]] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            cleaned = {str(k): "" if v is None else str(v) for k, v in row.items() if k}
            keys = candidate_manifest_keys(cleaned)
            for key in keys:
                manifest.setdefault(key, cleaned)
    return manifest


def candidate_manifest_keys(row: dict[str, str]) -> list[str]:
    key_names = [
        "file_name",
        "filename",
        "file",
        "path",
        "source_file",
        "workbook",
        "name",
    ]
    keys: list[str] = []
    for key_name in key_names:
        value = row.get(key_name)
        if not value:
            continue
        keys.append(value)
        keys.append(Path(value).name)
    return list(dict.fromkeys(keys))


def discover_workbooks(input_path: Path) -> list[Path]:
    if input_path.is_file():
        return [input_path] if input_path.suffix.lower() in KNOWN_WORKBOOK_EXTENSIONS else []
    return sorted(
        path
        for path in input_path.rglob("*")
        if path.is_file()
        and not path.name.startswith("~$")
        and path.suffix.lower() in KNOWN_WORKBOOK_EXTENSIONS
    )


def match_header_category(label: str) -> tuple[str, str] | None:
    normalized = normalize_label(label)
    if not normalized:
        return None
    for category, patterns in HEADER_PATTERNS.items():
        for pattern in patterns:
            normalized_pattern = normalize_label(pattern)
            if normalized == normalized_pattern or normalized_pattern in normalized:
                return category, pattern
    return None


def build_column_candidate(column_index: int, header: Any, matched_term: str) -> ColumnCandidate:
    return ColumnCandidate(
        column_index=column_index,
        column_letter=get_column_letter(column_index) if get_column_letter else str(column_index),
        header=str(header).strip(),
        confidence=1.0 if normalize_label(header) == normalize_label(matched_term) else 0.85,
        matched_term=matched_term,
    )


def score_header_row(values: list[Any]) -> tuple[float, list[str], list[str]]:
    categories: list[str] = []
    labels: list[str] = []
    non_empty = 0
    for value in values:
        label = normalize_label(value)
        if not label:
            continue
        non_empty += 1
        match = match_header_category(label)
        if match:
            category, _term = match
            categories.append(category)
            labels.append(str(value).strip())

    unique_categories = sorted(set(categories))
    score = len(unique_categories) * 2.0
    if "price" in unique_categories:
        score += 2.0
    if {"description", "price"}.issubset(unique_categories):
        score += 1.0
    if {"distributor_sku", "manufacturer_part_number"}.intersection(unique_categories):
        score += 1.0
    if non_empty >= 4:
        score += 0.5
    return score, unique_categories, labels


def used_range_for_sheet(sheet: Any) -> dict[str, int | str | None]:
    min_row: int | None = None
    max_row: int | None = None
    min_col: int | None = None
    max_col: int | None = None

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            min_row = cell.row if min_row is None else min(min_row, cell.row)
            max_row = cell.row if max_row is None else max(max_row, cell.row)
            min_col = cell.column if min_col is None else min(min_col, cell.column)
            max_col = cell.column if max_col is None else max(max_col, cell.column)

    if min_row is None or max_row is None or min_col is None or max_col is None:
        return {
            "min_row": None,
            "max_row": None,
            "min_column": None,
            "max_column": None,
            "range": None,
        }

    min_letter = get_column_letter(min_col) if get_column_letter else str(min_col)
    max_letter = get_column_letter(max_col) if get_column_letter else str(max_col)
    return {
        "min_row": min_row,
        "max_row": max_row,
        "min_column": min_col,
        "max_column": max_col,
        "range": f"{min_letter}{min_row}:{max_letter}{max_row}",
    }


def sheet_formula_count(sheet: Any) -> int:
    count = 0
    for row in sheet.iter_rows():
        for cell in row:
            if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("=")):
                count += 1
    return count


def hidden_count(dimensions: Iterable[Any]) -> int:
    return sum(1 for dimension in dimensions if getattr(dimension, "hidden", False))


def header_rows_for_sheet(sheet: Any) -> list[CandidateHeaderRow]:
    candidates: list[CandidateHeaderRow] = []
    max_row = min(sheet.max_row or 0, DEFAULT_HEADER_SCAN_ROWS)
    for row_index in range(1, max_row + 1):
        values = [cell.value for cell in sheet[row_index]]
        non_empty_count = sum(1 for value in values if normalize_label(value))
        score, categories, labels = score_header_row(values)
        if score >= 3.0:
            candidates.append(
                CandidateHeaderRow(
                    sheet_name=sheet.title,
                    row_index=row_index,
                    score=score,
                    non_empty_count=non_empty_count,
                    matched_categories=categories,
                    labels=labels,
                )
            )
    return sorted(candidates, key=lambda candidate: (-candidate.score, candidate.row_index))


def column_candidates_for_header(sheet: Any, header_row: int) -> dict[str, list[ColumnCandidate]]:
    candidates = {
        "distributor_sku": [],
        "manufacturer_part_number": [],
        "description": [],
        "uom": [],
        "price": [],
        "effective_date": [],
        "expiration_date": [],
    }
    for column_index, cell in enumerate(sheet[header_row], start=1):
        if cell.value is None:
            continue
        match = match_header_category(str(cell.value))
        if not match:
            continue
        category, matched_term = match
        candidates[category].append(build_column_candidate(column_index, cell.value, matched_term))
    return candidates


def redact_sample_value(value: Any, column_index: int, price_columns: set[int]) -> Any:
    if column_index in price_columns and value not in (None, ""):
        return "<redacted_price_value>"
    return display_cell_value(value)


def sample_rows_for_sheet(sheet: Any, header_row: int | None, price_columns: set[int]) -> list[dict[str, Any]]:
    if header_row is None:
        start_row = 1
    else:
        start_row = header_row + 1

    samples: list[dict[str, Any]] = []
    max_col = min(sheet.max_column or 0, DEFAULT_SAMPLE_COLUMNS)
    for row_index in range(start_row, (sheet.max_row or 0) + 1):
        row_values = {
            get_column_letter(column_index) if get_column_letter else str(column_index): redact_sample_value(
                sheet.cell(row=row_index, column=column_index).value,
                column_index,
                price_columns,
            )
            for column_index in range(1, max_col + 1)
        }
        if any(value not in (None, "") for value in row_values.values()):
            samples.append({"row_index": row_index, "values": row_values})
        if len(samples) >= DEFAULT_SAMPLE_ROWS:
            break
    return samples


def profile_sheet(sheet: Any) -> SheetProfile:
    candidates = header_rows_for_sheet(sheet)
    best_header = candidates[0] if candidates else None
    column_candidates = (
        column_candidates_for_header(sheet, best_header.row_index) if best_header else {}
    )
    price_columns = {
        candidate.column_index for candidate in column_candidates.get("price", [])
    }

    reasons: list[str] = []
    pricing_score = 0.0
    if best_header:
        pricing_score += best_header.score
        reasons.append(f"header row {best_header.row_index} scored {best_header.score:g}")
    if column_candidates.get("price"):
        pricing_score += 3.0
        reasons.append("price-like columns detected")
    if column_candidates.get("description"):
        pricing_score += 1.0
        reasons.append("description-like columns detected")
    if column_candidates.get("distributor_sku") or column_candidates.get("manufacturer_part_number"):
        pricing_score += 1.0
        reasons.append("part identifier columns detected")

    formula_count = sheet_formula_count(sheet)
    merged_ranges = [str(range_ref) for range_ref in sheet.merged_cells.ranges]
    row_hidden_count = hidden_count(sheet.row_dimensions.values())
    column_hidden_count = hidden_count(sheet.column_dimensions.values())
    risk_flags: list[str] = []
    if sheet.sheet_state != "visible":
        risk_flags.append("sheet_hidden")
    if formula_count:
        risk_flags.append("contains_formulas")
    if merged_ranges:
        risk_flags.append("contains_merged_cells")
    if row_hidden_count:
        risk_flags.append("contains_hidden_rows")
    if column_hidden_count:
        risk_flags.append("contains_hidden_columns")
    if not candidates:
        risk_flags.append("no_candidate_header_row")

    detected_labels = sorted(
        {label for candidate in candidates for label in candidate.labels}
    )

    return SheetProfile(
        name=sheet.title,
        state=sheet.sheet_state,
        dimensions={"max_row": sheet.max_row or 0, "max_column": sheet.max_column or 0},
        used_range=used_range_for_sheet(sheet),
        merged_cell_ranges=merged_ranges,
        formula_count=formula_count,
        hidden_row_count=row_hidden_count,
        hidden_column_count=column_hidden_count,
        candidate_header_rows=candidates,
        candidate_pricing_sheet_score=pricing_score,
        candidate_pricing_sheet_reasons=reasons,
        detected_header_labels=detected_labels,
        likely_distributor_sku_columns=column_candidates.get("distributor_sku", []),
        likely_manufacturer_part_number_columns=column_candidates.get(
            "manufacturer_part_number", []
        ),
        likely_description_columns=column_candidates.get("description", []),
        likely_uom_columns=column_candidates.get("uom", []),
        likely_price_columns=column_candidates.get("price", []),
        likely_effective_date_columns=column_candidates.get("effective_date", []),
        likely_expiration_date_columns=column_candidates.get("expiration_date", []),
        sample_rows=sample_rows_for_sheet(
            sheet,
            best_header.row_index if best_header else None,
            price_columns,
        ),
        risk_flags=risk_flags,
    )


def profile_workbook(path: Path, manifest: dict[str, dict[str, str]]) -> WorkbookProfile:
    extension = path.suffix.lower()
    profile = WorkbookProfile(
        file_name=path.name,
        file_hash=sha256_file(path),
        file_size=path.stat().st_size,
        workbook_extension=extension,
        manifest_metadata=manifest.get(path.name),
    )

    if extension == ".xls":
        profile.risk_flags.append("xls_not_supported_without_xlrd_adapter")
        return profile

    if extension not in SUPPORTED_OPENPYXL_EXTENSIONS:
        profile.risk_flags.append("unsupported_workbook_extension")
        return profile

    if load_workbook is None:
        profile.risk_flags.append("openpyxl_not_installed")
        return profile

    workbook = load_workbook(path, data_only=False, read_only=False)
    try:
        profile.sheet_names = list(workbook.sheetnames)
        profile.hidden_sheets = [
            sheet.title for sheet in workbook.worksheets if sheet.sheet_state != "visible"
        ]
        profile.sheets = [profile_sheet(sheet) for sheet in workbook.worksheets]
    finally:
        workbook.close()

    profile.candidate_pricing_sheets = [
        {
            "sheet_name": sheet.name,
            "score": sheet.candidate_pricing_sheet_score,
            "reasons": sheet.candidate_pricing_sheet_reasons,
        }
        for sheet in sorted(
            profile.sheets,
            key=lambda sheet: (-sheet.candidate_pricing_sheet_score, sheet.name),
        )
        if sheet.candidate_pricing_sheet_score >= 5
    ]

    if not profile.candidate_pricing_sheets:
        profile.risk_flags.append("no_candidate_pricing_sheet")
    if profile.hidden_sheets:
        profile.risk_flags.append("contains_hidden_sheets")
    if any(sheet.formula_count for sheet in profile.sheets):
        profile.risk_flags.append("contains_formulas")
    if len(profile.candidate_pricing_sheets) > 1:
        profile.risk_flags.append("multiple_candidate_pricing_sheets")

    return profile


def safe_profile_file_name(path: Path) -> str:
    safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", path.name)
    return f"{safe_name}.pricing_discovery.json"


def write_json_profile(profile: WorkbookProfile, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / safe_profile_file_name(Path(profile.file_name))
    output_path.write_text(
        json.dumps(asdict(profile), indent=2, sort_keys=True),
        encoding="utf-8",
    )
    return output_path


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def flatten_workbook_inventory(profiles: list[WorkbookProfile]) -> list[dict[str, Any]]:
    return [
        {
            "file_name": profile.file_name,
            "file_hash": profile.file_hash,
            "file_size": profile.file_size,
            "workbook_extension": profile.workbook_extension,
            "sheet_count": len(profile.sheet_names),
            "hidden_sheet_count": len(profile.hidden_sheets),
            "candidate_pricing_sheet_count": len(profile.candidate_pricing_sheets),
            "risk_flags": ";".join(profile.risk_flags),
        }
        for profile in profiles
    ]


def flatten_sheet_inventory(profiles: list[WorkbookProfile]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for profile in profiles:
        for sheet in profile.sheets:
            rows.append(
                {
                    "file_name": profile.file_name,
                    "sheet_name": sheet.name,
                    "state": sheet.state,
                    "max_row": sheet.dimensions["max_row"],
                    "max_column": sheet.dimensions["max_column"],
                    "used_range": sheet.used_range["range"],
                    "merged_cell_range_count": len(sheet.merged_cell_ranges),
                    "formula_count": sheet.formula_count,
                    "hidden_row_count": sheet.hidden_row_count,
                    "hidden_column_count": sheet.hidden_column_count,
                    "candidate_pricing_sheet_score": sheet.candidate_pricing_sheet_score,
                    "risk_flags": ";".join(sheet.risk_flags),
                }
            )
    return rows


def flatten_candidate_headers(profiles: list[WorkbookProfile]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for profile in profiles:
        for sheet in profile.sheets:
            for candidate in sheet.candidate_header_rows:
                rows.append(
                    {
                        "file_name": profile.file_name,
                        "sheet_name": sheet.name,
                        "row_index": candidate.row_index,
                        "score": candidate.score,
                        "non_empty_count": candidate.non_empty_count,
                        "matched_categories": ";".join(candidate.matched_categories),
                        "labels": " | ".join(candidate.labels),
                    }
                )
    return rows


def write_reports(profiles: list[WorkbookProfile], output_dir: Path) -> None:
    write_csv(
        output_dir / "workbook_inventory.csv",
        [
            "file_name",
            "file_hash",
            "file_size",
            "workbook_extension",
            "sheet_count",
            "hidden_sheet_count",
            "candidate_pricing_sheet_count",
            "risk_flags",
        ],
        flatten_workbook_inventory(profiles),
    )
    write_csv(
        output_dir / "sheet_inventory.csv",
        [
            "file_name",
            "sheet_name",
            "state",
            "max_row",
            "max_column",
            "used_range",
            "merged_cell_range_count",
            "formula_count",
            "hidden_row_count",
            "hidden_column_count",
            "candidate_pricing_sheet_score",
            "risk_flags",
        ],
        flatten_sheet_inventory(profiles),
    )
    write_csv(
        output_dir / "candidate_headers.csv",
        [
            "file_name",
            "sheet_name",
            "row_index",
            "score",
            "non_empty_count",
            "matched_categories",
            "labels",
        ],
        flatten_candidate_headers(profiles),
    )
    write_risk_report(profiles, output_dir / "risk_report.md")
    write_discovery_summary(profiles, output_dir / "discovery_summary.md")


def write_risk_report(profiles: list[WorkbookProfile], path: Path) -> None:
    lines = ["# Pricing Discovery Risk Report", ""]
    any_risks = False
    for profile in profiles:
        sheet_risks = [
            f"{sheet.name}: {', '.join(sheet.risk_flags)}"
            for sheet in profile.sheets
            if sheet.risk_flags
        ]
        if not profile.risk_flags and not sheet_risks:
            continue
        any_risks = True
        lines.append(f"## {profile.file_name}")
        if profile.risk_flags:
            lines.append(f"- Workbook: {', '.join(profile.risk_flags)}")
        for sheet_risk in sheet_risks:
            lines.append(f"- Sheet: {sheet_risk}")
        lines.append("")
    if not any_risks:
        lines.append("No structural risk flags were detected.")
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def write_discovery_summary(profiles: list[WorkbookProfile], path: Path) -> None:
    workbook_count = len(profiles)
    sheet_count = sum(len(profile.sheets) for profile in profiles)
    candidate_count = sum(len(profile.candidate_pricing_sheets) for profile in profiles)
    risky_count = sum(
        1
        for profile in profiles
        if profile.risk_flags or any(sheet.risk_flags for sheet in profile.sheets)
    )
    lines = [
        "# Pricing Discovery Summary",
        "",
        f"- Workbooks profiled: {workbook_count}",
        f"- Sheets profiled: {sheet_count}",
        f"- Candidate pricing sheets: {candidate_count}",
        f"- Workbooks with risk flags: {risky_count}",
        "",
        "This discovery phase records workbook structure and candidate mappings only.",
        "It does not import prices into Prometheus, call the OpenAI API, or write to Zeus.",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run_discovery(input_path: Path, manifest_path: Path | None, output_dir: Path) -> list[WorkbookProfile]:
    manifest = load_manifest(manifest_path)
    workbooks = discover_workbooks(input_path)
    profiles = [profile_workbook(path, manifest) for path in workbooks]

    profiles_dir = output_dir / "profiles"
    for profile in profiles:
        write_json_profile(profile, profiles_dir)
    write_reports(profiles, output_dir)
    return profiles


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Discover deterministic structure metadata for contract-pricing workbooks."
    )
    parser.add_argument("--input", required=True, type=Path, help="Workbook file or folder.")
    parser.add_argument(
        "--manifest",
        type=Path,
        default=None,
        help="Optional manifest CSV with workbook metadata.",
    )
    parser.add_argument("--output", required=True, type=Path, help="Output directory.")
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_arg_parser()
    args = parser.parse_args(argv)
    profiles = run_discovery(args.input, args.manifest, args.output)
    print(
        json.dumps(
            {
                "workbooks_profiled": len(profiles),
                "output": str(args.output),
            },
            indent=2,
            sort_keys=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

