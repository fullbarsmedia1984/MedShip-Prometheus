from __future__ import annotations

import csv
import hashlib
import json
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from pricing_ingestion.normalization import apply_transform, normalize_currency, normalize_uom, parse_currency, parse_date
from pricing_ingestion.profiles.loader import ITEM_IDENTIFIER_FIELDS, load_profile, validate_profile


PROPOSED_ROW_FIELDS = [
    "ingestion_row_id",
    "profile_name",
    "profile_version",
    "distributor_name",
    "distributor_id",
    "contract_name",
    "contract_number",
    "account_number",
    "location",
    "internal_item_id",
    "distributor_sku",
    "model_number",
    "manufacturer_name",
    "manufacturer_part_number",
    "gtin",
    "udi",
    "ndc",
    "item_description_raw",
    "item_description_normalized",
    "raw_uom",
    "normalized_uom",
    "raw_price_uom",
    "normalized_price_uom",
    "raw_base_uom",
    "normalized_base_uom",
    "raw_pack_size",
    "pack_size",
    "raw_price",
    "price",
    "currency",
    "tier",
    "effective_date",
    "expiration_date",
    "minimum_quantity",
    "rebate_terms",
    "freight_terms",
    "source_file",
    "source_file_hash",
    "source_sheet_name",
    "source_sheet_hidden",
    "source_row_number",
    "source_row_hidden",
    "source_column_map",
    "source_cell_map",
    "metadata_source_map",
    "formula_fields",
    "validation_status",
    "exception_codes",
    "warning_codes",
]

EXCEPTION_FIELDS = [
    "exception_code",
    "severity",
    "source_file",
    "source_sheet",
    "source_row",
    "source_cell",
    "canonical_field",
    "raw_value_summary",
    "message",
]

EXCLUDED_ROW_FIELDS = [
    "source_file",
    "source_sheet",
    "source_row",
    "reason",
    "message",
    "decision_id",
]

METADATA_FIELDS = {
    "contract_name",
    "contract_number",
    "account_number",
    "location",
    "effective_date",
    "expiration_date",
}


@dataclass
class DryRunException:
    code: str
    severity: str
    source_file: str
    source_sheet: str
    source_row: int | None
    source_cell: str | None
    canonical_field: str | None
    raw_value_summary: str
    message: str


@dataclass
class DryRunResult:
    proposed_rows: list[dict[str, Any]] = field(default_factory=list)
    exceptions: list[DryRunException] = field(default_factory=list)
    excluded_rows: list[dict[str, Any]] = field(default_factory=list)
    rows_scanned: int = 0
    repeated_header_rows: int = 0
    total_rows: int = 0
    duplicate_rows: int = 0
    formula_derived_rows: int = 0
    hidden_content_rows: int = 0


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_manifest(path: Path | None) -> dict[str, dict[str, str]]:
    if path is None or not path.exists():
        return {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return {row.get("file_name", ""): row for row in csv.DictReader(handle)}


def normalize_header(value: Any) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"[_\-/#:]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip().lower()


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def redacted_summary(value: Any) -> str:
    if value is None:
        return "blank"
    text = str(value).strip()
    if not text:
        return "blank"
    if len(text) > 40:
        return f"text_length={len(text)}"
    if re.search(r"\d", text):
        return f"value_length={len(text)}"
    return text


def json_default(value: Any) -> Any:
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)


def csv_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (dict, list, set, tuple)):
        return json.dumps(value, sort_keys=True, default=json_default)
    return str(value)


def matches_pattern(value: str, pattern: str) -> bool:
    return bool(re.search(pattern, value, re.IGNORECASE))


def select_sheets(workbook: Any, profile: dict[str, Any]) -> list[Any]:
    rules = profile.get("sheet_rules", {})
    selected = []
    for sheet in workbook.worksheets:
        if sheet.sheet_state != "visible" and not rules.get("include_hidden_sheets", False):
            continue
        name = sheet.title
        include = True
        if rules.get("exact_sheet_name"):
            include = name == rules["exact_sheet_name"]
        if rules.get("case_insensitive_sheet_name"):
            include = name.lower() == str(rules["case_insensitive_sheet_name"]).lower()
        if rules.get("regex_sheet_name"):
            include = matches_pattern(name, rules["regex_sheet_name"])
        include_patterns = rules.get("include_patterns", [])
        if include_patterns:
            include = any(matches_pattern(name, pattern) for pattern in include_patterns)
        exclude_patterns = rules.get("exclude_patterns", [])
        if any(matches_pattern(name, pattern) for pattern in exclude_patterns):
            include = False
        if include:
            selected.append(sheet)
    if not rules.get("process_multiple_matching_sheets", False):
        return selected[:1]
    return selected


def resolve_header_row(sheet: Any, profile: dict[str, Any]) -> tuple[int, dict[str, int]]:
    rules = profile.get("header_rules", {})
    if rules.get("exact_header_row"):
        row_index = int(rules["exact_header_row"])
        return row_index, header_map_for_row(sheet, row_index)

    row_range = rules.get("candidate_row_range", {"start": 1, "end": min(sheet.max_row or 1, 50)})
    required_terms = [normalize_header(value) for value in rules.get("required_header_terms", [])]
    best_row = None
    best_score = -1
    best_map: dict[str, int] = {}
    for row_index in range(int(row_range["start"]), int(row_range["end"]) + 1):
        candidate_map = header_map_for_row(sheet, row_index)
        headers = set(candidate_map)
        score = sum(1 for term in required_terms if term in headers)
        if score > best_score:
            best_score = score
            best_row = row_index
            best_map = candidate_map
    if best_row is None or best_score < rules.get("minimum_header_match_score", 0):
        raise ValueError("Header row could not be detected")
    return best_row, best_map


def header_map_for_row(sheet: Any, row_index: int) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for cell in sheet[row_index]:
        normalized = normalize_header(cell.value)
        if normalized:
            header_map.setdefault(normalized, cell.column)
    return header_map


def mapping_column(mapping: dict[str, Any], header_map: dict[str, int]) -> int | None:
    columns = mapping_columns(mapping, header_map)
    return columns[0] if columns else None


def mapping_columns(mapping: dict[str, Any], header_map: dict[str, int]) -> list[int]:
    columns: list[int] = []
    if mapping.get("source_column_index"):
        columns.append(int(mapping["source_column_index"]))
    elif mapping.get("source_column_letter"):
        letters = str(mapping["source_column_letter"]).upper()
        column = 0
        for char in letters:
            column = column * 26 + (ord(char) - ord("A") + 1)
        columns.append(column)
    candidates = []
    if not columns and mapping.get("source_header"):
        candidates.append(mapping["source_header"])
    if not columns:
        candidates.extend(mapping.get("source_header_aliases", []))
    candidates.extend(mapping.get("fallback_source_headers", []))
    for candidate in candidates:
        normalized = normalize_header(candidate)
        if normalized in header_map and header_map[normalized] not in columns:
            columns.append(header_map[normalized])
    return columns


def is_repeated_header(row_values: list[Any], header_map: dict[str, int]) -> bool:
    headers = set(header_map)
    row_headers = {normalize_header(value) for value in row_values if normalize_header(value)}
    return bool(row_headers) and len(row_headers & headers) >= max(2, min(3, len(headers)))


def is_total_row(row_values: list[Any]) -> bool:
    labels = " ".join(str(value).strip().lower() for value in row_values if value is not None)
    return any(term in labels for term in ["subtotal", "grand total", "summary", " total"])


def has_item_identifier(row: dict[str, Any]) -> bool:
    return any(row.get(field) for field in ITEM_IDENTIFIER_FIELDS)


def cell_reference(row: int, column: int) -> str:
    return f"{get_column_letter(column)}{row}"


def transform_value(raw_value: Any, mapping: dict[str, Any]) -> tuple[Any, str | None]:
    value = raw_value
    for transform in mapping.get("transforms", []):
        try:
            value = apply_transform(value, transform)
        except ValueError as exc:
            return None, str(exc)
    return value, None


def field_raw_name(field: str) -> str | None:
    if field == "price":
        return "raw_price"
    if field in {"raw_uom", "raw_price_uom", "raw_base_uom"}:
        return field
    if field == "pack_size":
        return "raw_pack_size"
    return None


def add_exception(result: DryRunResult, code: str, severity: str, source_file: str, source_sheet: str, source_row: int | None, source_cell: str | None, field: str | None, raw_value: Any, message: str) -> None:
    result.exceptions.append(
        DryRunException(
            code=code,
            severity=severity,
            source_file=source_file,
            source_sheet=source_sheet,
            source_row=source_row,
            source_cell=source_cell,
            canonical_field=field,
            raw_value_summary=redacted_summary(raw_value),
            message=message,
        )
    )


def metadata_missing_code(field: str) -> str:
    if field == "effective_date":
        return "MISSING_REQUIRED_EFFECTIVE_DATE"
    return "MISSING_REQUIRED_CONTRACT_METADATA"


def apply_default_values(row: dict[str, Any], profile: dict[str, Any], manifest_row: dict[str, str], result: DryRunResult, source_file: str, sheet_name: str, row_number: int) -> None:
    for field, config in profile.get("default_values", {}).items():
        if row.get(field) not in (None, ""):
            row.setdefault("metadata_source_map", {})[field] = "workbook"
            continue
        source = config.get("source")
        if config.get("source") == "constant":
            row[field] = config.get("value")
            row.setdefault("metadata_source_map", {})[field] = "profile_default"
        elif config.get("source") == "manifest":
            manifest_field = config.get("manifest_field")
            value = manifest_row.get(manifest_field or "", "")
            row[field] = value or None
            row.setdefault("metadata_source_map", {})[field] = "manifest" if value else "manifest_missing"
            if config.get("required") and not value:
                row.setdefault("exception_codes", []).append(metadata_missing_code(field))
                add_exception(
                    result,
                    metadata_missing_code(field),
                    "blocking",
                    source_file,
                    sheet_name,
                    row_number,
                    None,
                    field,
                    None,
                    f"Required metadata missing for {field}.",
                )
        elif config.get("required"):
            row.setdefault("metadata_source_map", {})[field] = source or "unapproved"
            row.setdefault("exception_codes", []).append("METADATA_SOURCE_UNAPPROVED")
            add_exception(
                result,
                "METADATA_SOURCE_UNAPPROVED",
                "blocking",
                source_file,
                sheet_name,
                row_number,
                None,
                field,
                None,
                f"Metadata source is not approved for {field}.",
            )


def validate_row(row: dict[str, Any], profile: dict[str, Any], result: DryRunResult, source_file: str, sheet_name: str, row_number: int) -> tuple[list[str], list[str]]:
    exceptions: list[str] = []
    warnings: list[str] = []

    if row.get("price") in (None, ""):
        exceptions.append("MISSING_PRICE")
    else:
        try:
            price = parse_currency(row.get("price"))
            row["price"] = price
            if price is None:
                exceptions.append("MISSING_PRICE")
            elif price <= 0:
                exceptions.append("NON_POSITIVE_PRICE")
        except ValueError:
            exceptions.append("INVALID_PRICE")

    if not row.get("item_description_raw"):
        exceptions.append("MISSING_DESCRIPTION")
    if not any(row.get(field) for field in ITEM_IDENTIFIER_FIELDS):
        exceptions.append("MISSING_ITEM_IDENTIFIER")

    if not row.get("raw_uom"):
        warnings.append("MISSING_UOM")
    elif row.get("normalized_uom") in (None, "") and "UNKNOWN_UOM" not in row.get("warning_codes", []):
        try:
            row["normalized_uom"] = normalize_uom(row.get("raw_uom"))
        except ValueError:
            warnings.append("UNKNOWN_UOM")

    for raw_field, normalized_field in [
        ("raw_price_uom", "normalized_price_uom"),
        ("raw_base_uom", "normalized_base_uom"),
    ]:
        if row.get(raw_field) and row.get(normalized_field) in (None, ""):
            try:
                row[normalized_field] = normalize_uom(row.get(raw_field))
            except ValueError:
                if "UNKNOWN_UOM" not in row.get("warning_codes", []) and "UNKNOWN_UOM" not in warnings:
                    warnings.append("UNKNOWN_UOM")

    for source_field, target_field, code in [
        ("effective_date", "effective_date", "INVALID_EFFECTIVE_DATE"),
        ("expiration_date", "expiration_date", "INVALID_EXPIRATION_DATE"),
    ]:
        if row.get(source_field):
            try:
                row[target_field] = parse_date(row.get(source_field))
            except ValueError:
                exceptions.append(code)

    if row.get("effective_date") and row.get("expiration_date"):
        if str(row["expiration_date"]) < str(row["effective_date"]):
            exceptions.append("EXPIRATION_BEFORE_EFFECTIVE_DATE")

    for field in profile.get("validation_rules", {}).get("required_metadata_fields", []):
        if row.get(field) in (None, ""):
            code = metadata_missing_code(str(field))
            if code not in row.get("exception_codes", []):
                exceptions.append(code)

    for code in exceptions:
        add_exception(result, code, "blocking", source_file, sheet_name, row_number, None, None, None, code.replace("_", " ").title())
    for code in warnings:
        add_exception(result, code, "warning", source_file, sheet_name, row_number, None, None, None, code.replace("_", " ").title())
    return exceptions, warnings


def duplicate_key(row: dict[str, Any]) -> tuple[Any, ...]:
    return (
        row.get("distributor_sku"),
        row.get("manufacturer_part_number"),
        row.get("gtin"),
        row.get("udi"),
        row.get("ndc"),
        row.get("item_description_normalized"),
        row.get("normalized_uom"),
        row.get("minimum_quantity"),
    )


def parse_sheet(
    workbook_formula: Any,
    workbook_values: Any,
    sheet: Any,
    profile: dict[str, Any],
    workbook_path: Path,
    file_hash: str,
    manifest_row: dict[str, str],
    result: DryRunResult,
) -> None:
    sheet_values = workbook_values[sheet.title]
    header_row, header_map = resolve_header_row(sheet, profile)
    data_rules = profile.get("data_region_rules", {})
    first_row = int(data_rules.get("explicit_first_data_row") or (header_row + 1))
    if data_rules.get("first_row_after_detected_header", False):
        first_row = header_row + 1
    last_row = int(data_rules.get("explicit_last_row") or (sheet.max_row or first_row))
    if data_rules.get("maximum_row_limit"):
        last_row = min(last_row, first_row + int(data_rules["maximum_row_limit"]) - 1)

    column_lookup = {mapping["canonical_field"]: mapping_column(mapping, header_map) for mapping in profile.get("column_mappings", [])}
    column_candidates = {mapping["canonical_field"]: mapping_columns(mapping, header_map) for mapping in profile.get("column_mappings", [])}
    for mapping in profile.get("column_mappings", []):
        if mapping.get("required") and column_lookup.get(mapping["canonical_field"]) is None:
            add_exception(
                result,
                "MISSING_REQUIRED_COLUMN",
                "blocking",
                workbook_path.name,
                sheet.title,
                header_row,
                None,
                mapping["canonical_field"],
                None,
                f"Required source column not found for {mapping['canonical_field']}.",
            )

    blank_required_rows = 0
    seen_duplicates: dict[tuple[Any, ...], dict[str, Any]] = {}
    for row_number in range(first_row, last_row + 1):
        row_values = [cell.value for cell in sheet[row_number]]
        if data_rules.get("repeated_header_detection", True) and is_repeated_header(row_values, header_map):
            result.repeated_header_rows += 1
            result.excluded_rows.append({"source_file": workbook_path.name, "source_sheet": sheet.title, "source_row": row_number, "reason": "REPEATED_HEADER_SKIPPED", "message": "Repeated header row skipped.", "decision_id": ""})
            continue
        if data_rules.get("skip_subtotal_and_total_rows", True) and is_total_row(row_values):
            result.total_rows += 1
            result.excluded_rows.append({"source_file": workbook_path.name, "source_sheet": sheet.title, "source_row": row_number, "reason": "TOTAL_ROW_SKIPPED", "message": "Subtotal/total row skipped.", "decision_id": ""})
            continue
        if sheet.row_dimensions[row_number].hidden and data_rules.get("skip_hidden_rows", False):
            result.excluded_rows.append({"source_file": workbook_path.name, "source_sheet": sheet.title, "source_row": row_number, "reason": "HIDDEN_ROW_SKIPPED", "message": "Hidden row skipped by profile rule.", "decision_id": ""})
            continue

        result.rows_scanned += 1
        proposed: dict[str, Any] = {field: None for field in PROPOSED_ROW_FIELDS}
        proposed.update(
            {
                "ingestion_row_id": f"{profile['profile_name']}:{workbook_path.name}:{sheet.title}:{row_number}",
                "profile_name": profile["profile_name"],
                "profile_version": profile["profile_version"],
                "distributor_name": profile["distributor_name"],
                "distributor_id": profile.get("distributor_id"),
                "internal_item_id": None,
                "source_file": workbook_path.name,
                "source_file_hash": file_hash,
                "source_sheet_name": sheet.title,
                "source_sheet_hidden": sheet.sheet_state != "visible",
                "source_row_number": row_number,
                "source_row_hidden": bool(sheet.row_dimensions[row_number].hidden),
                "source_column_map": {},
                "source_cell_map": {},
                "metadata_source_map": {},
                "formula_fields": [],
                "exception_codes": [],
                "warning_codes": [],
            }
        )

        required_blanks = 0
        mapped_required = 0
        formula_warning = False
        hidden_warning = proposed["source_sheet_hidden"] or proposed["source_row_hidden"]

        for mapping in profile.get("column_mappings", []):
            field = mapping["canonical_field"]
            column = column_lookup.get(field)
            if mapping.get("constant_value") is not None:
                raw_value = mapping.get("constant_value")
                source_cell = None
            elif not column_candidates.get(field):
                continue
            else:
                raw_value = None
                formula_cell = None
                value_cell = None
                chosen_column = column
                for candidate_column in column_candidates[field]:
                    candidate_formula_cell = sheet.cell(row=row_number, column=candidate_column)
                    candidate_value_cell = sheet_values.cell(row=row_number, column=candidate_column)
                    candidate_raw_value = candidate_value_cell.value if candidate_formula_cell.data_type == "f" else candidate_formula_cell.value
                    if raw_value is None or not is_blank(candidate_raw_value):
                        raw_value = candidate_raw_value
                        formula_cell = candidate_formula_cell
                        value_cell = candidate_value_cell
                        chosen_column = candidate_column
                    if not is_blank(candidate_raw_value):
                        break
                if formula_cell is None:
                    continue
                source_cell = cell_reference(row_number, chosen_column)
                proposed["source_column_map"][field] = get_column_letter(chosen_column)
                proposed["source_cell_map"][field] = source_cell
                if sheet.column_dimensions[get_column_letter(chosen_column)].hidden:
                    hidden_warning = True
                if isinstance(formula_cell, MergedCell):
                    add_exception(result, "AMBIGUOUS_MERGED_CELL", "blocking", workbook_path.name, sheet.title, row_number, source_cell, field, raw_value, "Mapped cell is merged and ambiguous.")
                if formula_cell.data_type == "f" or (isinstance(formula_cell.value, str) and formula_cell.value.startswith("=")):
                    if raw_value is None:
                        add_exception(result, "FORMULA_WITHOUT_CACHED_VALUE", "blocking", workbook_path.name, sheet.title, row_number, source_cell, field, raw_value, "Formula cell has no cached value.")
                    else:
                        formula_warning = True
                        proposed["formula_fields"].append(field)
            raw_field = field_raw_name(field)
            if raw_field:
                proposed[raw_field] = raw_value

            value, error = transform_value(raw_value, mapping)
            if error:
                if field in {"raw_uom", "raw_price_uom", "raw_base_uom"}:
                    if "UNKNOWN_UOM" not in proposed["warning_codes"]:
                        proposed["warning_codes"].append("UNKNOWN_UOM")
                        add_exception(result, "UNKNOWN_UOM", "warning", workbook_path.name, sheet.title, row_number, source_cell, field, raw_value, error)
                elif field == "pack_size":
                    proposed["warning_codes"].append("PACK_SIZE_NOT_PARSED")
                    add_exception(result, "PACK_SIZE_NOT_PARSED", "warning", workbook_path.name, sheet.title, row_number, source_cell, field, raw_value, error)
                elif field == "effective_date":
                    add_exception(result, "INVALID_EFFECTIVE_DATE", "blocking", workbook_path.name, sheet.title, row_number, source_cell, field, raw_value, error)
                    proposed["exception_codes"].append("INVALID_EFFECTIVE_DATE")
                elif field == "expiration_date":
                    add_exception(result, "INVALID_EXPIRATION_DATE", "blocking", workbook_path.name, sheet.title, row_number, source_cell, field, raw_value, error)
                    proposed["exception_codes"].append("INVALID_EXPIRATION_DATE")
                else:
                    code = "INVALID_PRICE" if field == "price" else "INVALID_PROFILE_MAPPING"
                    add_exception(result, code, "blocking", workbook_path.name, sheet.title, row_number, source_cell, field, raw_value, error)
                    proposed["exception_codes"].append(code)
            if field == "item_description_raw":
                proposed["item_description_raw"] = raw_value
                proposed["item_description_normalized"] = value
            elif field == "raw_uom":
                proposed["raw_uom"] = raw_value
                proposed["normalized_uom"] = value
            elif field == "raw_price_uom":
                proposed["raw_price_uom"] = raw_value
                proposed["normalized_price_uom"] = value
            elif field == "raw_base_uom":
                proposed["raw_base_uom"] = raw_value
                proposed["normalized_base_uom"] = value
            elif field == "pack_size":
                proposed["raw_pack_size"] = raw_value
                proposed["pack_size"] = value
            else:
                if field in METADATA_FIELDS and proposed.get(field) not in (None, "") and value not in (None, "") and proposed.get(field) != value:
                    proposed["exception_codes"].append("METADATA_CONFLICT")
                    add_exception(result, "METADATA_CONFLICT", "blocking", workbook_path.name, sheet.title, row_number, source_cell, field, None, "Metadata value conflicts with another approved source.")
                proposed[field] = value
                if field in METADATA_FIELDS and value not in (None, ""):
                    proposed["metadata_source_map"][field] = "workbook"

            if mapping.get("required"):
                mapped_required += 1
                if is_blank(raw_value):
                    required_blanks += 1

        if mapped_required and required_blanks == mapped_required:
            blank_required_rows += 1
            if blank_required_rows >= data_rules.get("stop_after_consecutive_blank_required_field_rows", 5):
                break
            continue
        blank_required_rows = 0

        if not has_item_identifier(proposed) and proposed.get("price") in (None, ""):
            result.excluded_rows.append(
                {
                    "source_file": workbook_path.name,
                    "source_sheet": sheet.title,
                    "source_row": row_number,
                    "reason": "SECTION_OR_NON_ITEM_ROW_SKIPPED",
                    "message": "Row lacks item identifier and price; treated as non-item structural row.",
                    "decision_id": "",
                }
            )
            continue

        apply_default_values(proposed, profile, manifest_row, result, workbook_path.name, sheet.title, row_number)

        if formula_warning:
            result.formula_derived_rows += 1
            proposed["warning_codes"].append("FORMULA_DERIVED_PRICE")
            add_exception(result, "FORMULA_DERIVED_PRICE", "warning", workbook_path.name, sheet.title, row_number, None, None, None, "Formula-derived mapped value used.")
        if hidden_warning:
            result.hidden_content_rows += 1
            code = "HIDDEN_SHEET" if proposed["source_sheet_hidden"] else "HIDDEN_ROW"
            proposed["warning_codes"].append(code)
            add_exception(result, code, "warning", workbook_path.name, sheet.title, row_number, None, None, None, "Hidden source content used.")

        exception_codes, warning_codes = validate_row(proposed, profile, result, workbook_path.name, sheet.title, row_number)
        proposed["exception_codes"].extend(exception_codes)
        proposed["warning_codes"].extend(warning_codes)

        key = duplicate_key(proposed)
        if key in seen_duplicates:
            result.duplicate_rows += 1
            prior = seen_duplicates[key]
            if prior.get("price") == proposed.get("price"):
                proposed["warning_codes"].append("DUPLICATE_IDENTICAL_ROW")
                add_exception(result, "DUPLICATE_IDENTICAL_ROW", "warning", workbook_path.name, sheet.title, row_number, None, None, None, "Duplicate identical row detected.")
            else:
                proposed["exception_codes"].append("DUPLICATE_CONFLICTING_PRICE")
                add_exception(result, "DUPLICATE_CONFLICTING_PRICE", "blocking", workbook_path.name, sheet.title, row_number, None, None, None, "Duplicate key has conflicting price.")
        else:
            seen_duplicates[key] = proposed

        proposed["validation_status"] = "blocking" if proposed["exception_codes"] else ("warning" if proposed["warning_codes"] else "valid")
        result.proposed_rows.append(proposed)


def run_dry_run(workbook_path: Path, profile_path: Path, manifest_path: Path | None, output_dir: Path) -> dict[str, Any]:
    profile = load_profile(profile_path)
    validation_errors = validate_profile(profile)
    if validation_errors:
        raise ValueError(f"Profile validation failed: {[error.__dict__ for error in validation_errors]}")

    manifest = load_manifest(manifest_path)
    manifest_row = manifest.get(workbook_path.name, {})
    file_hash = sha256_file(workbook_path)
    result = DryRunResult()

    workbook_formula = load_workbook(workbook_path, data_only=False, read_only=False)
    workbook_values = load_workbook(workbook_path, data_only=True, read_only=False)
    try:
        selected_sheets = select_sheets(workbook_formula, profile)
        if not selected_sheets:
            add_exception(result, "INVALID_PROFILE_MAPPING", "blocking", workbook_path.name, "", None, None, None, None, "No sheet matched profile rules.")
        for sheet in selected_sheets:
            parse_sheet(workbook_formula, workbook_values, sheet, profile, workbook_path, file_hash, manifest_row, result)
    finally:
        workbook_formula.close()
        workbook_values.close()

    output_dir.mkdir(parents=True, exist_ok=True)
    write_outputs(result, output_dir)
    summary = summarize(result, output_dir)
    return summary


def summarize(result: DryRunResult, output_dir: Path) -> dict[str, Any]:
    valid_rows = sum(1 for row in result.proposed_rows if row["validation_status"] == "valid")
    warning_rows = sum(1 for row in result.proposed_rows if row["validation_status"] == "warning")
    blocking_rows = sum(1 for row in result.proposed_rows if row["validation_status"] == "blocking")
    exception_counts = {}
    for exception in result.exceptions:
        exception_counts[exception.code] = exception_counts.get(exception.code, 0) + 1
    summary = {
        "rows_scanned": result.rows_scanned,
        "proposed_rows": len(result.proposed_rows),
        "valid_rows": valid_rows,
        "warning_rows": warning_rows,
        "blocking_exception_rows": blocking_rows,
        "excluded_header_rows": result.repeated_header_rows,
        "excluded_total_rows": result.total_rows,
        "duplicate_rows": result.duplicate_rows,
        "formula_derived_rows": result.formula_derived_rows,
        "hidden_content_rows": result.hidden_content_rows,
        "exception_counts": exception_counts,
        "output": str(output_dir),
    }
    (output_dir / "dry_run_summary.json").write_text(json.dumps(summary, indent=2, sort_keys=True), encoding="utf-8")
    return summary


def write_outputs(result: DryRunResult, output_dir: Path) -> None:
    with (output_dir / "proposed_rows.csv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=PROPOSED_ROW_FIELDS, extrasaction="ignore")
        writer.writeheader()
        for row in result.proposed_rows:
            writer.writerow({field: csv_value(row.get(field)) for field in PROPOSED_ROW_FIELDS})

    with (output_dir / "exceptions.csv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=EXCEPTION_FIELDS)
        writer.writeheader()
        for exception in result.exceptions:
            writer.writerow(
                {
                    "exception_code": exception.code,
                    "severity": exception.severity,
                    "source_file": exception.source_file,
                    "source_sheet": exception.source_sheet,
                    "source_row": exception.source_row or "",
                    "source_cell": exception.source_cell or "",
                    "canonical_field": exception.canonical_field or "",
                    "raw_value_summary": exception.raw_value_summary,
                    "message": exception.message,
                }
            )

    with (output_dir / "excluded_rows.csv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=EXCLUDED_ROW_FIELDS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(result.excluded_rows)

    summary = summarize(result, output_dir)
    write_summary_markdown(summary, output_dir / "dry_run_summary.md")
    write_mapping_review(result, output_dir / "mapping_review.md")


def write_summary_markdown(summary: dict[str, Any], path: Path) -> None:
    lines = [
        "# Contract Pricing Dry Run Summary",
        "",
        f"- Rows scanned: {summary['rows_scanned']}",
        f"- Proposed rows: {summary['proposed_rows']}",
        f"- Valid rows: {summary['valid_rows']}",
        f"- Warning rows: {summary['warning_rows']}",
        f"- Blocking-exception rows: {summary['blocking_exception_rows']}",
        f"- Excluded repeated header rows: {summary['excluded_header_rows']}",
        f"- Excluded total/subtotal rows: {summary['excluded_total_rows']}",
        f"- Duplicate rows: {summary['duplicate_rows']}",
        f"- Formula-derived rows: {summary['formula_derived_rows']}",
        f"- Hidden-content rows: {summary['hidden_content_rows']}",
        "",
        "## Exception Counts",
        "",
    ]
    if summary["exception_counts"]:
        lines.extend(f"- {code}: {count}" for code, count in sorted(summary["exception_counts"].items()))
    else:
        lines.append("- None")
    lines.append("")
    lines.append("This report intentionally excludes raw prices and full source rows.")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_mapping_review(result: DryRunResult, path: Path) -> None:
    mappings: dict[str, tuple[str, str]] = {}
    for row in result.proposed_rows:
        source_column_map = row.get("source_column_map") or {}
        source_cell_map = row.get("source_cell_map") or {}
        if isinstance(source_column_map, dict):
            for field, column in source_column_map.items():
                mappings.setdefault(field, (str(column), str(source_cell_map.get(field, ""))))
    lines = [
        "# Mapping Review",
        "",
        "| Canonical field | Source column | Example source cell | Review notes |",
        "| --- | --- | --- | --- |",
    ]
    for field, (column, source_cell) in sorted(mappings.items()):
        lines.append(f"| {field} | {column} | {source_cell} | Confirm mapping and transforms. |")
    if not mappings:
        lines.append("| None |  |  | No proposed row mappings were produced. |")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
