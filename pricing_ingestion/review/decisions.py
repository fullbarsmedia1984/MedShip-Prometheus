from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from pricing_ingestion.dry_run.runner import is_blank, mapping_columns, normalize_header, resolve_header_row
from pricing_ingestion.profiles.loader import ITEM_IDENTIFIER_FIELDS, load_profile
from pricing_ingestion.review.exceptions import DryRunReview, load_review_runs


BUSINESS_REVIEW_OUTPUT = Path("outputs/pricing_discovery/business_review")
UOM_DECISION_PATH = Path("data/pricing/uom_decisions.local.csv")
RAW_ROOT = Path("data/pricing/raw")

PROFILE_REQUIRED_CATEGORIES = [
    "pricing_basis",
    "item_identifier",
    "price_uom",
    "base_uom",
    "contract_metadata",
    "effective_date_source",
    "row_exclusion_policy",
]

METADATA_POLICY_IDS = [
    "contract_number_required_before_publish",
    "account_number_required_before_publish",
    "location_required_before_publish",
    "effective_date_required_before_publish",
    "expiration_date_nullable",
    "metadata_source_policy",
]

DECISION_STATUSES = {"pending", "approved", "rejected", "not_applicable"}
TOP_LEVEL_STATUSES = {"draft", "incomplete", "ready_to_apply", "applied", "superseded"}

UOM_DECISION_FIELDS = [
    "profile_name",
    "raw_uom",
    "normalized_token",
    "occurrence_count",
    "classification",
    "proposed_normalized_uom",
    "selected_normalized_uom",
    "decision_status",
    "rationale",
    "decided_by",
    "decided_at",
]

PRICE_PRESENCE_FIELDS = [
    "profile_name",
    "workbook_variant",
    "sheet_name",
    "selected_price_header",
    "alternate_price_header",
    "selected_price_present",
    "alternate_price_present",
    "identifier_present",
    "description_present",
    "formula_present",
    "row_classification",
    "occurrence_count",
    "recommended_review_action",
]

MANUAL_ROW_LOCATION_FIELDS = [
    "workbook",
    "sheet",
    "row_number",
    "classification",
    "selected_price_present",
    "alternate_price_present",
    "identifier_present",
    "description_present",
]

PRICE_BASIS_FIELDS = [
    "profile_name",
    "family_id",
    "source_header",
    "structural_classification",
    "populated_row_count",
    "blank_row_count",
    "formula_row_count",
    "relative_column_position",
    "candidate_role",
    "recommendation",
    "confidence",
    "requires_business_confirmation",
    "notes",
]

IDENTIFIER_FIELDS = [
    "profile_name",
    "family_id",
    "source_header",
    "candidate_canonical_field",
    "occurrence_count",
    "uniqueness_rate",
    "blank_rate",
    "formatting_pattern_summary",
    "competing_identifier_headers",
    "recommendation",
    "confidence",
    "requires_business_confirmation",
    "notes",
]

METADATA_FIELDS = [
    "profile_name",
    "workbook_count",
    "distributor_name_complete",
    "contract_name_complete",
    "contract_number_complete",
    "account_number_complete",
    "location_complete",
    "effective_date_complete",
    "expiration_date_complete",
    "required_before_publish",
    "notes",
]

DUPLICATE_FIELDS = [
    "profile_name",
    "workbook_variant",
    "duplicate_identical_count",
    "duplicate_conflicting_price_count",
    "recommended_review_action",
]


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def selected_runs(runs: list[DryRunReview]) -> list[DryRunReview]:
    hardening = [run for run in runs if run.run_name.startswith("hardening_v1_")]
    if not hardening:
        return runs
    covered_sources = {tuple(run.source_files) for run in hardening}
    extras = [run for run in runs if tuple(run.source_files) not in covered_sources and not run.run_name.startswith("hardening_v1_")]
    return sorted(hardening + extras, key=lambda run: run.run_name)


def family_id(profile: dict[str, Any]) -> str:
    return ";".join(str(value) for value in profile.get("applicable_family_ids", []))


def mapping_for(profile: dict[str, Any], canonical_field: str) -> dict[str, Any] | None:
    for mapping in profile.get("column_mappings", []):
        if mapping.get("canonical_field") == canonical_field:
            return mapping
    return None


def mapping_header(profile: dict[str, Any], canonical_field: str) -> str:
    mapping = mapping_for(profile, canonical_field)
    if not mapping:
        return ""
    return str(mapping.get("source_header") or mapping.get("source_column_letter") or "")


def candidate_price_headers(profile: dict[str, Any], workbook_path: Path | None = None) -> list[str]:
    headers: list[str] = []
    mapped = mapping_header(profile, "price")
    if mapped:
        headers.append(mapped)
    for rule_header in profile.get("workbook_match_rules", {}).get("optional_headers", []):
        if re.search(r"price|cost|discount|net|customer", str(rule_header), re.IGNORECASE):
            headers.append(str(rule_header))
    if workbook_path and workbook_path.exists():
        workbook = load_workbook(workbook_path, data_only=False, read_only=False)
        try:
            for sheet in workbook.worksheets:
                if not sheet_matches_profile(sheet.title, profile):
                    continue
                header_row, _ = resolve_header_row(sheet, profile)
                for cell in sheet[header_row]:
                    if cell.value and re.search(r"price|cost|discount|net|customer", str(cell.value), re.IGNORECASE):
                        headers.append(str(cell.value).strip())
        finally:
            workbook.close()
    return sorted(dict.fromkeys(header for header in headers if header))


def sheet_matches_profile(sheet_name: str, profile: dict[str, Any]) -> bool:
    rules = profile.get("sheet_rules", {})
    if rules.get("exact_sheet_name"):
        return sheet_name == rules["exact_sheet_name"]
    if rules.get("case_insensitive_sheet_name"):
        return sheet_name.lower() == str(rules["case_insensitive_sheet_name"]).lower()
    if rules.get("regex_sheet_name"):
        return bool(re.search(str(rules["regex_sheet_name"]), sheet_name, re.IGNORECASE))
    patterns = rules.get("include_patterns", [])
    if patterns:
        return any(re.search(str(pattern), sheet_name, re.IGNORECASE) for pattern in patterns)
    return True


def workbook_path_for(run: DryRunReview, raw_root: Path) -> Path | None:
    if not run.source_files:
        return None
    path = raw_root / run.source_files[0]
    return path if path.exists() else None


def yes_no(value: bool | None) -> str:
    if value is None:
        return "unknown"
    return "yes" if value else "no"


def detecto_missing_price_locations(run: DryRunReview) -> set[int]:
    rows = set()
    for exception in run.exceptions:
        if exception.get("exception_code") == "MISSING_PRICE" and str(exception.get("source_row", "")).isdigit():
            rows.add(int(exception["source_row"]))
    return rows


def proposed_by_row(run: DryRunReview) -> dict[int, dict[str, str]]:
    by_row = {}
    for row in run.proposed_rows:
        source_row = str(row.get("source_row_number") or "")
        if source_row.isdigit():
            by_row[int(source_row)] = row
    return by_row


def classify_price_presence(
    selected_present: bool | None,
    alternate_present: bool | None,
    identifier_present: bool | None,
    description_present: bool | None,
    formula_present: bool,
    duplicate_related: bool,
) -> str:
    if duplicate_related:
        return "duplicate_related"
    if formula_present and not selected_present:
        return "formula_without_cached_value"
    if selected_present is False and alternate_present is True:
        return "selected_price_blank_alternate_present"
    if selected_present is False and alternate_present is False:
        if identifier_present and description_present:
            return "identifier_and_description_present_no_price"
        if not identifier_present and not description_present:
            return "likely_non_item_row"
        return "all_candidate_prices_blank"
    if identifier_present is False and description_present is False:
        return "likely_non_item_row"
    if identifier_present != description_present:
        return "malformed_row"
    return "unknown"


def review_action_for_classification(classification: str) -> str:
    actions = {
        "selected_price_blank_alternate_present": "Confirm whether alternate price field is list-only or contract basis.",
        "all_candidate_prices_blank": "Review whether rows should stay blocking or be excluded by policy.",
        "formula_without_cached_value": "Refresh workbook or request value-only export before publish.",
        "likely_non_item_row": "Confirm row-exclusion policy before suppressing.",
        "identifier_and_description_present_no_price": "Keep blocking until pricing policy is approved.",
        "malformed_row": "Review workbook structure and mapping assumptions.",
        "duplicate_related": "Review duplicate handling and conflict policy.",
        "unknown": "Manual review required.",
    }
    return actions[classification]


def build_detecto_price_presence_rows(runs: list[DryRunReview], raw_root: Path = RAW_ROOT) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    aggregate = Counter()
    locations: list[dict[str, Any]] = []
    for run in runs:
        if not run.profile_name.startswith("detecto"):
            continue
        missing_rows = detecto_missing_price_locations(run)
        if not missing_rows:
            continue
        proposed_rows = proposed_by_row(run)
        duplicate_rows = {
            int(exception["source_row"])
            for exception in run.exceptions
            if exception.get("exception_code") in {"DUPLICATE_IDENTICAL_ROW", "DUPLICATE_CONFLICTING_PRICE"}
            and str(exception.get("source_row", "")).isdigit()
        }
        selected_header = mapping_header(run.profile, "price")
        alternate_headers = [header for header in candidate_price_headers(run.profile, workbook_path_for(run, raw_root)) if header != selected_header]
        alternate_header = alternate_headers[0] if alternate_headers else ""
        workbook_path = workbook_path_for(run, raw_root)

        raw_states = inspect_row_presence(workbook_path, run.profile, sorted(missing_rows), selected_header, alternate_header)
        for row_number in sorted(missing_rows):
            proposed = proposed_rows.get(row_number, {})
            state = raw_states.get(row_number, {})
            selected_present = state.get("selected_price_present")
            alternate_present = state.get("alternate_price_present")
            identifier_present = state.get("identifier_present")
            if identifier_present is None:
                identifier_present = any(proposed.get(field) for field in ITEM_IDENTIFIER_FIELDS)
            description_present = state.get("description_present")
            if description_present is None:
                description_present = bool(proposed.get("item_description_raw"))
            formula_present = bool(state.get("formula_present")) or "FORMULA_DERIVED_PRICE" in str(proposed.get("warning_codes", ""))
            classification = classify_price_presence(
                selected_present,
                alternate_present,
                identifier_present,
                description_present,
                formula_present,
                row_number in duplicate_rows,
            )
            key = (
                run.profile_name,
                run.run_name,
                state.get("sheet_name") or proposed.get("source_sheet_name") or "",
                selected_header,
                alternate_header,
                yes_no(selected_present),
                yes_no(alternate_present),
                yes_no(identifier_present),
                yes_no(description_present),
                yes_no(formula_present),
                classification,
                review_action_for_classification(classification),
            )
            aggregate[key] += 1
            locations.append(
                {
                    "workbook": run.source_files[0] if run.source_files else "",
                    "sheet": state.get("sheet_name") or proposed.get("source_sheet_name") or "",
                    "row_number": row_number,
                    "classification": classification,
                    "selected_price_present": yes_no(selected_present),
                    "alternate_price_present": yes_no(alternate_present),
                    "identifier_present": yes_no(identifier_present),
                    "description_present": yes_no(description_present),
                }
            )
    rows = [
        {
            "profile_name": key[0],
            "workbook_variant": key[1],
            "sheet_name": key[2],
            "selected_price_header": key[3],
            "alternate_price_header": key[4],
            "selected_price_present": key[5],
            "alternate_price_present": key[6],
            "identifier_present": key[7],
            "description_present": key[8],
            "formula_present": key[9],
            "row_classification": key[10],
            "occurrence_count": count,
            "recommended_review_action": key[11],
        }
        for key, count in sorted(aggregate.items())
    ]
    return rows, locations


def inspect_row_presence(workbook_path: Path | None, profile: dict[str, Any], row_numbers: list[int], selected_header: str, alternate_header: str) -> dict[int, dict[str, Any]]:
    if not workbook_path or not workbook_path.exists() or not row_numbers:
        return {}
    formula_wb = load_workbook(workbook_path, data_only=False, read_only=False)
    value_wb = load_workbook(workbook_path, data_only=True, read_only=False)
    states: dict[int, dict[str, Any]] = {}
    try:
        for sheet in formula_wb.worksheets:
            if not sheet_matches_profile(sheet.title, profile):
                continue
            value_sheet = value_wb[sheet.title]
            header_row, header_map = resolve_header_row(sheet, profile)
            selected_col = header_map.get(normalize_header(selected_header))
            alternate_col = header_map.get(normalize_header(alternate_header)) if alternate_header else None
            identifier_cols = []
            for field in ITEM_IDENTIFIER_FIELDS:
                mapping = mapping_for(profile, field)
                if mapping:
                    identifier_cols.extend(mapping_columns(mapping, header_map))
            desc_mapping = mapping_for(profile, "item_description_raw")
            desc_cols = mapping_columns(desc_mapping, header_map) if desc_mapping else []
            for row_number in row_numbers:
                selected_cell = sheet.cell(row=row_number, column=selected_col) if selected_col else None
                selected_value_cell = value_sheet.cell(row=row_number, column=selected_col) if selected_col else None
                alternate_cell = sheet.cell(row=row_number, column=alternate_col) if alternate_col else None
                alternate_value_cell = value_sheet.cell(row=row_number, column=alternate_col) if alternate_col else None
                formula_present = any(
                    cell is not None and (cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("=")))
                    for cell in [selected_cell, alternate_cell]
                )
                states[row_number] = {
                    "sheet_name": sheet.title,
                    "selected_price_present": None if selected_cell is None else not is_blank(selected_value_cell.value if selected_cell.data_type == "f" else selected_cell.value),
                    "alternate_price_present": None if alternate_cell is None else not is_blank(alternate_value_cell.value if alternate_cell.data_type == "f" else alternate_cell.value),
                    "identifier_present": any(not is_blank(sheet.cell(row=row_number, column=column).value) for column in identifier_cols),
                    "description_present": any(not is_blank(sheet.cell(row=row_number, column=column).value) for column in desc_cols),
                    "formula_present": formula_present,
                }
    finally:
        formula_wb.close()
        value_wb.close()
    return states


def price_role_for_header(header: str) -> str:
    normalized = header.lower()
    if "net" in normalized:
        return "net_price"
    if "customer" in normalized:
        return "customer_price"
    if "discount" in normalized:
        return "discounted_price"
    if "list" in normalized:
        return "list_price"
    if "promo" in normalized:
        return "promotional_price"
    if "cost" in normalized:
        return "cost"
    if "contract" in normalized:
        return "contract_price"
    return "unknown"


def build_price_basis_candidates(runs: list[DryRunReview], raw_root: Path = RAW_ROOT) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for run in runs:
        workbook_path = workbook_path_for(run, raw_root)
        headers = candidate_price_headers(run.profile, workbook_path)
        counts = inspect_column_counts(workbook_path, run.profile, headers)
        for position, header in enumerate(headers, start=1):
            if (run.profile_name, header) in seen:
                continue
            seen.add((run.profile_name, header))
            role = price_role_for_header(header)
            if role == "list_price":
                recommendation = "Requires explicit business confirmation before using as contract pricing."
                confidence = "high"
            elif role in {"net_price", "customer_price", "discounted_price", "contract_price"}:
                recommendation = "Candidate pricing basis; confirm negotiated contract meaning."
                confidence = "medium"
            else:
                recommendation = "Unknown pricing role; do not approve without business confirmation."
                confidence = "low"
            count = counts.get(header, {})
            rows.append(
                {
                    "profile_name": run.profile_name,
                    "family_id": run.family_id,
                    "source_header": header,
                    "structural_classification": "price_like_header",
                    "populated_row_count": count.get("populated", 0),
                    "blank_row_count": count.get("blank", 0),
                    "formula_row_count": count.get("formula", 0),
                    "relative_column_position": position,
                    "candidate_role": role,
                    "recommendation": recommendation,
                    "confidence": confidence,
                    "requires_business_confirmation": "yes",
                    "notes": "No negotiated-price conclusion is inferred from numeric population alone.",
                }
            )
    return rows


def inspect_column_counts(workbook_path: Path | None, profile: dict[str, Any], headers: list[str]) -> dict[str, dict[str, int]]:
    counts = {header: {"populated": 0, "blank": 0, "formula": 0} for header in headers}
    if not workbook_path or not workbook_path.exists():
        return counts
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        for sheet in workbook.worksheets:
            if not sheet_matches_profile(sheet.title, profile):
                continue
            header_row, header_map = resolve_header_row(sheet, profile)
            first_row = int(profile.get("data_region_rules", {}).get("explicit_first_data_row") or header_row + 1)
            last_row = int(profile.get("data_region_rules", {}).get("explicit_last_row") or sheet.max_row or first_row)
            for header in headers:
                column = header_map.get(normalize_header(header))
                if not column:
                    continue
                for row_number in range(first_row, last_row + 1):
                    cell = sheet.cell(row=row_number, column=column)
                    if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("=")):
                        counts[header]["formula"] += 1
                    if is_blank(cell.value):
                        counts[header]["blank"] += 1
                    else:
                        counts[header]["populated"] += 1
    finally:
        workbook.close()
    return counts


def identifier_headers(profile: dict[str, Any], workbook_path: Path | None = None) -> list[str]:
    headers = []
    for mapping in profile.get("column_mappings", []):
        if mapping.get("canonical_field") in ITEM_IDENTIFIER_FIELDS:
            headers.append(str(mapping.get("source_header") or ""))
    for header in profile.get("workbook_match_rules", {}).get("required_headers", []) + profile.get("workbook_match_rules", {}).get("optional_headers", []):
        if re.search(r"sku|item|model|part|upc|gtin|udi|ndc|number|#", str(header), re.IGNORECASE):
            headers.append(str(header))
    if workbook_path and workbook_path.exists():
        workbook = load_workbook(workbook_path, data_only=False, read_only=False)
        try:
            for sheet in workbook.worksheets:
                if not sheet_matches_profile(sheet.title, profile):
                    continue
                header_row, _ = resolve_header_row(sheet, profile)
                for cell in sheet[header_row]:
                    if cell.value and re.search(r"sku|item|model|part|upc|gtin|udi|ndc|number|#", str(cell.value), re.IGNORECASE):
                        headers.append(str(cell.value).strip())
        finally:
            workbook.close()
    return sorted(dict.fromkeys(header for header in headers if header))


def canonical_candidates_for_identifier(profile_name: str, header: str, mapped_field: str | None) -> list[str]:
    normalized = header.lower()
    if profile_name.startswith("detecto") and "model" in normalized:
        return ["manufacturer_part_number", "distributor_sku", "model_number"]
    if mapped_field:
        return [mapped_field]
    if "upc" in normalized or "gtin" in normalized:
        return ["gtin"]
    if "sku" in normalized or "item" in normalized or "#" in normalized:
        return ["distributor_sku"]
    if "model" in normalized:
        return ["model_number"]
    if "part" in normalized:
        return ["manufacturer_part_number"]
    return ["model_number"]


def build_identifier_candidates(runs: list[DryRunReview], raw_root: Path = RAW_ROOT) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str]] = set()
    for run in runs:
        workbook_path = workbook_path_for(run, raw_root)
        headers = identifier_headers(run.profile, workbook_path)
        counts = inspect_identifier_counts(workbook_path, run.profile, headers)
        competing = "; ".join(headers)
        mapped_by_header = {
            str(mapping.get("source_header") or ""): str(mapping.get("canonical_field") or "")
            for mapping in run.profile.get("column_mappings", [])
        }
        for header in headers:
            for candidate in canonical_candidates_for_identifier(run.profile_name, header, mapped_by_header.get(header)):
                key = (run.profile_name, header, candidate)
                if key in seen:
                    continue
                seen.add(key)
                stats = counts.get(header, {"total": 0, "blank": 0, "unique": 0, "patterns": Counter()})
                total = int(stats["total"]) or 0
                blank = int(stats["blank"]) or 0
                unique = int(stats["unique"]) or 0
                uniqueness_rate = f"{(unique / max(total - blank, 1) * 100):.1f}%"
                blank_rate = f"{(blank / max(total, 1) * 100):.1f}%"
                pattern_summary = "; ".join(f"{name}:{count}" for name, count in sorted(stats["patterns"].items())) or "unknown"
                recommend_model = run.profile_name.startswith("detecto") and header.lower() == "model" and candidate == "model_number"
                rows.append(
                    {
                        "profile_name": run.profile_name,
                        "family_id": run.family_id,
                        "source_header": header,
                        "candidate_canonical_field": candidate,
                        "occurrence_count": max(total - blank, 0),
                        "uniqueness_rate": uniqueness_rate,
                        "blank_rate": blank_rate,
                        "formatting_pattern_summary": pattern_summary,
                        "competing_identifier_headers": competing,
                        "recommendation": "Recommend model_number unless business confirms a stronger SKU/MPN meaning." if recommend_model else "Requires business confirmation.",
                        "confidence": "medium" if recommend_model else "low",
                        "requires_business_confirmation": "yes",
                        "notes": "Identifier values are not included in this aggregate report.",
                    }
                )
    return rows


def pattern_for_identifier(value: Any) -> str:
    text = str(value or "")
    if not text:
        return "blank"
    if re.fullmatch(r"\d+", text):
        return "numeric"
    if re.fullmatch(r"[A-Za-z]+", text):
        return "alpha"
    if re.fullmatch(r"[A-Za-z0-9]+", text):
        return "alphanumeric"
    return "mixed_format"


def inspect_identifier_counts(workbook_path: Path | None, profile: dict[str, Any], headers: list[str]) -> dict[str, dict[str, Any]]:
    counts: dict[str, dict[str, Any]] = {
        header: {"total": 0, "blank": 0, "unique_values": set(), "patterns": Counter()} for header in headers
    }
    if not workbook_path or not workbook_path.exists():
        return {header: {"total": 0, "blank": 0, "unique": 0, "patterns": Counter()} for header in headers}
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    try:
        for sheet in workbook.worksheets:
            if not sheet_matches_profile(sheet.title, profile):
                continue
            header_row, header_map = resolve_header_row(sheet, profile)
            first_row = int(profile.get("data_region_rules", {}).get("explicit_first_data_row") or header_row + 1)
            last_row = int(profile.get("data_region_rules", {}).get("explicit_last_row") or sheet.max_row or first_row)
            for header in headers:
                column = header_map.get(normalize_header(header))
                if not column:
                    continue
                for row_number in range(first_row, last_row + 1):
                    value = sheet.cell(row=row_number, column=column).value
                    counts[header]["total"] += 1
                    if is_blank(value):
                        counts[header]["blank"] += 1
                    else:
                        counts[header]["unique_values"].add(str(value))
                    counts[header]["patterns"][pattern_for_identifier(value)] += 1
    finally:
        workbook.close()
    return {
        header: {
            "total": int(row["total"]),
            "blank": int(row["blank"]),
            "unique": len(row["unique_values"]),
            "patterns": row["patterns"],
        }
        for header, row in counts.items()
    }


def build_metadata_completeness(runs: list[DryRunReview], manifest_path: Path) -> list[dict[str, Any]]:
    manifest_rows = {row.get("file_name", ""): row for row in read_csv(manifest_path)}
    rows = []
    for profile_name, profile_runs in group_runs_by_profile(runs).items():
        source_files = sorted({source for run in profile_runs for source in run.source_files})
        field_presence = {}
        for field in ["distributor_name", "contract_name", "contract_number", "account_number", "location", "effective_date", "expiration_date"]:
            present_count = sum(1 for source in source_files if manifest_rows.get(source, {}).get(field))
            if not source_files:
                field_presence[field] = "unknown"
            elif present_count == len(source_files):
                field_presence[field] = "yes"
            elif present_count == 0:
                field_presence[field] = "no"
            else:
                field_presence[field] = "partial"
        rows.append(
            {
                "profile_name": profile_name,
                "workbook_count": len(source_files),
                "distributor_name_complete": field_presence["distributor_name"],
                "contract_name_complete": field_presence["contract_name"],
                "contract_number_complete": field_presence["contract_number"],
                "account_number_complete": field_presence["account_number"],
                "location_complete": field_presence["location"],
                "effective_date_complete": field_presence["effective_date"],
                "expiration_date_complete": field_presence["expiration_date"],
                "required_before_publish": "pending metadata policy decision",
                "notes": "Presence/absence only; metadata values are intentionally excluded.",
            }
        )
    return rows


def build_duplicate_review(runs: list[DryRunReview]) -> list[dict[str, Any]]:
    rows = []
    for run in runs:
        counts = Counter(exception.get("exception_code", "") for exception in run.exceptions)
        rows.append(
            {
                "profile_name": run.profile_name,
                "workbook_variant": run.run_name,
                "duplicate_identical_count": counts.get("DUPLICATE_IDENTICAL_ROW", 0),
                "duplicate_conflicting_price_count": counts.get("DUPLICATE_CONFLICTING_PRICE", 0),
                "recommended_review_action": "Review duplicate policy before approval." if counts.get("DUPLICATE_CONFLICTING_PRICE", 0) else "No blocking duplicate action required unless business policy requires it.",
            }
        )
    return rows


def group_runs_by_profile(runs: list[DryRunReview]) -> dict[str, list[DryRunReview]]:
    grouped: dict[str, list[DryRunReview]] = defaultdict(list)
    for run in runs:
        grouped[run.profile_name].append(run)
    return dict(grouped)


def normalize_review_uom_classification(classification: str) -> str:
    mapping = {
        "known_alias": "direct_lexical_synonym",
        "direct_lexical_synonym": "direct_lexical_synonym",
        "packaging_expression": "packaging_expression",
        "dimensional_unit": "dimensional_unit",
        "ambiguous_unit": "ambiguous_unit",
        "missing_value": "missing_value",
        "invalid_value": "invalid_value",
        "requires_business_review": "ambiguous_unit",
    }
    return mapping.get(classification, "ambiguous_unit")


def build_uom_decisions(review_root: Path) -> list[dict[str, Any]]:
    rows = []
    for row in read_csv(review_root / "uom_review.csv"):
        classification = normalize_review_uom_classification(row.get("classification", ""))
        candidate = row.get("candidate_normalized_uom", "") if classification == "direct_lexical_synonym" else ""
        rows.append(
            {
                "profile_name": row.get("profile_name", ""),
                "raw_uom": row.get("raw_uom", ""),
                "normalized_token": row.get("normalized_token", ""),
                "occurrence_count": int(row.get("occurrence_count") or 0),
                "classification": classification,
                "candidate_normalized_uom": candidate,
                "proposed_normalized_uom": candidate,
                "selected_normalized_uom": "",
                "decision_status": "pending",
                "rationale": "",
                "decided_by": "",
                "decided_at": "",
            }
        )
    return rows


def decision(
    profile_name: str,
    family: str,
    decision_id: str,
    category: str,
    question: str,
    options: list[str],
    evidence_reference: str,
) -> dict[str, Any]:
    return {
        "profile_name": profile_name,
        "family_id": family,
        "decision_id": decision_id,
        "category": category,
        "question": question,
        "candidate_options": options,
        "selected_option": "",
        "rationale": "",
        "evidence_reference": evidence_reference,
        "decision_status": "pending",
        "decided_by": "",
        "decided_at": "",
    }


def build_profile_decisions(runs: list[DryRunReview], price_candidates: list[dict[str, Any]], identifier_candidates: list[dict[str, Any]]) -> list[dict[str, Any]]:
    price_options = options_by_profile(price_candidates, "source_header", "candidate_role")
    identifier_options = options_by_profile(identifier_candidates, "source_header", "candidate_canonical_field")
    decisions = []
    for profile_name, profile_runs in sorted(group_runs_by_profile(runs).items()):
        profile = profile_runs[0].profile
        family = family_id(profile)
        distributor = str(profile.get("distributor_name") or profile_name)
        decisions.extend(
            [
                decision(profile_name, family, f"{profile_name}:pricing_basis", "pricing_basis", f"For {distributor}, which source field is the approved contract pricing basis?", price_options.get(profile_name, ["unresolved"]), "outputs/pricing_discovery/business_review/price_basis_candidates.csv"),
                decision(profile_name, family, f"{profile_name}:item_identifier", "item_identifier", "What canonical meaning should the primary source identifier have?", identifier_options.get(profile_name, ["unresolved"]), "outputs/pricing_discovery/business_review/identifier_candidates.csv"),
                decision(profile_name, family, f"{profile_name}:price_uom", "price_uom", "Which source field defines the unit to which the price applies?", price_uom_options(profile), "outputs/pricing_discovery/business_review/uom_decision_template.csv"),
                decision(profile_name, family, f"{profile_name}:base_uom", "base_uom", "Should the canonical schema be extended with raw_price_uom, normalized_price_uom, raw_base_uom, and normalized_base_uom?", ["extend_schema_pending_publish", "do_not_extend_schema", "unresolved"], "outputs/pricing_discovery/business_review/business_review_packet.md"),
                decision(profile_name, family, f"{profile_name}:contract_metadata", "contract_metadata", "Which metadata source is authoritative for contract, account, location, and date fields?", ["manifest", "profile_default", "workbook", "explicit_upload_input", "unresolved"], "outputs/pricing_discovery/business_review/metadata_completeness.csv"),
                decision(profile_name, family, f"{profile_name}:effective_date_source", "effective_date_source", "Which source should populate effective and expiration dates before publish?", ["manifest", "workbook_columns", "explicit_upload_input", "may_be_blank_until_publish", "unresolved"], "outputs/pricing_discovery/business_review/metadata_completeness.csv"),
                decision(profile_name, family, f"{profile_name}:row_exclusion_policy", "row_exclusion_policy", "Which missing-price or structural rows may be excluded instead of blocking?", ["keep_identifier_description_no_price_blocking", "exclude_confirmed_non_item_rows_only", "exclude_all_blank_candidate_price_rows", "unresolved"], "outputs/pricing_discovery/business_review/detecto_price_presence.csv"),
            ]
        )
        if profile_name.startswith("detecto"):
            decisions.append(
                decision(profile_name, family, f"{profile_name}:profile_variant", "profile_variant", "Should Detecto keep one shared profile or split the larger workbook into a variant profile?", ["keep_single_profile", "split_larger_workbook_variant", "unresolved"], "outputs/pricing_discovery/business_review/detecto_price_presence.csv")
            )
    return decisions


def price_uom_options(profile: dict[str, Any]) -> list[str]:
    options = [mapping_header(profile, "raw_uom") or "UOM", "another_detected_field", "unresolved"]
    optional = profile.get("workbook_match_rules", {}).get("optional_headers", [])
    for header in optional:
        if re.search(r"uom|unit", str(header), re.IGNORECASE):
            options.insert(1, str(header))
    return sorted(dict.fromkeys(options), key=options.index)


def options_by_profile(rows: list[dict[str, Any]], label_field: str, detail_field: str) -> dict[str, list[str]]:
    options: dict[str, list[str]] = defaultdict(list)
    for row in rows:
        label = str(row.get(label_field) or "unresolved")
        detail = str(row.get(detail_field) or "unknown")
        option = f"{label} [{detail}]"
        if option not in options[str(row.get("profile_name") or "")]:
            options[str(row.get("profile_name") or "")].append(option)
    return dict(options)


def build_metadata_policy_decisions() -> list[dict[str, Any]]:
    questions = {
        "contract_number_required_before_publish": "Is contract_number required before publish?",
        "account_number_required_before_publish": "Is account_number required before publish?",
        "location_required_before_publish": "Is location required before publish?",
        "effective_date_required_before_publish": "Is effective_date required before publish?",
        "expiration_date_nullable": "May expiration_date be null?",
        "metadata_source_policy": "May metadata come from manifest, profile default, workbook, or explicit upload input?",
    }
    option_map = {
        "expiration_date_nullable": ["nullable_allowed", "required_before_publish", "unresolved"],
        "metadata_source_policy": ["manifest", "profile_default", "workbook", "explicit_upload_input", "unresolved"],
    }
    rows = []
    for decision_id in METADATA_POLICY_IDS:
        rows.append(
            decision(
                "global",
                "",
                f"metadata_policy:{decision_id}",
                "contract_metadata",
                questions[decision_id],
                option_map.get(decision_id, ["required_before_publish", "optional_before_publish", "unresolved"]),
                "outputs/pricing_discovery/business_review/metadata_completeness.csv",
            )
        )
    return rows


def merge_decisions(generated: list[dict[str, Any]], existing: list[dict[str, Any]]) -> list[dict[str, Any]]:
    existing_by_id = {row.get("decision_id"): row for row in existing}
    merged = []
    for row in generated:
        prior = existing_by_id.get(row["decision_id"])
        if prior:
            for field in ["selected_option", "rationale", "decision_status", "decided_by", "decided_at"]:
                row[field] = prior.get(field, row.get(field, ""))
        merged.append(row)
    return merged


def merge_uom_decisions(generated: list[dict[str, Any]], existing: list[dict[str, str]]) -> list[dict[str, Any]]:
    existing_by_key = {(row.get("profile_name"), row.get("raw_uom"), row.get("normalized_token")): row for row in existing}
    merged = []
    for row in generated:
        prior = existing_by_key.get((row["profile_name"], row["raw_uom"], row["normalized_token"]))
        if prior:
            for field in ["selected_normalized_uom", "decision_status", "rationale", "decided_by", "decided_at"]:
                row[field] = prior.get(field, row.get(field, ""))
        merged.append(row)
    return merged


def build_business_decisions(
    runs: list[DryRunReview],
    review_root: Path,
    price_candidates: list[dict[str, Any]],
    identifier_candidates: list[dict[str, Any]],
    existing: dict[str, Any] | None = None,
) -> dict[str, Any]:
    now = utc_now()
    profile_decisions = build_profile_decisions(runs, price_candidates, identifier_candidates)
    metadata_decisions = build_metadata_policy_decisions()
    uom_decisions = build_uom_decisions(review_root)
    if existing:
        profile_decisions = merge_decisions(profile_decisions, existing.get("profiles", []))
        metadata_decisions = merge_decisions(metadata_decisions, existing.get("metadata_policy", {}).get("decisions", []))
        uom_decisions = merge_uom_decisions(uom_decisions, existing.get("uom_decisions", []))
    return {
        "schema_version": "1.0.0",
        "decision_set_version": existing.get("decision_set_version") if existing else now,
        "status": existing.get("status", "draft") if existing else "draft",
        "created_at": existing.get("created_at") if existing else now,
        "updated_at": now,
        "decided_by": existing.get("decided_by") if existing else None,
        "profiles": profile_decisions,
        "uom_decisions": [{key: value for key, value in row.items() if key != "proposed_normalized_uom"} for row in uom_decisions],
        "metadata_policy": {
            "decisions": metadata_decisions,
            "notes": ["Metadata values are stored locally only; this decision file records policy choices."],
        },
        "notes": [
            "Generated for human review only.",
            "No business decisions are applied automatically by this file.",
            "No unit conversions or each prices are inferred.",
        ],
    }


def validate_decisions(data: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    if data.get("status") not in TOP_LEVEL_STATUSES:
        errors.append("INVALID_TOP_LEVEL_STATUS")
    profile_decisions = data.get("profiles", [])
    metadata_decisions = data.get("metadata_policy", {}).get("decisions", [])
    all_profile_decisions = profile_decisions + metadata_decisions
    seen_profile_category: set[tuple[str, str]] = set()
    categories_by_profile: dict[str, set[str]] = defaultdict(set)
    for row in all_profile_decisions:
        status = row.get("decision_status")
        if status not in DECISION_STATUSES:
            errors.append(f"INVALID_DECISION_STATUS:{row.get('decision_id')}")
        selected = row.get("selected_option") or ""
        if selected and selected not in row.get("candidate_options", []):
            errors.append(f"SELECTED_OPTION_NOT_IN_CANDIDATES:{row.get('decision_id')}")
        if status == "approved" and (not row.get("decided_by") or not row.get("decided_at")):
            errors.append(f"APPROVED_DECISION_REQUIRES_REVIEWER_AND_DATE:{row.get('decision_id')}")
        if status == "approved" and row.get("category") == "pricing_basis" and "list_price" in selected and not row.get("rationale"):
            errors.append(f"LIST_PRICE_SELECTION_REQUIRES_RATIONALE:{row.get('decision_id')}")
        key = (str(row.get("profile_name") or ""), str(row.get("category") or ""))
        if row.get("profile_name") != "global":
            if key in seen_profile_category:
                errors.append(f"CONFLICTING_PROFILE_CATEGORY_DECISIONS:{row.get('profile_name')}:{row.get('category')}")
            seen_profile_category.add(key)
            categories_by_profile[str(row.get("profile_name") or "")].add(str(row.get("category") or ""))
    for profile_name, categories in categories_by_profile.items():
        missing = set(PROFILE_REQUIRED_CATEGORIES) - categories
        for category in sorted(missing):
            errors.append(f"MISSING_REQUIRED_DECISION:{profile_name}:{category}")
    for decision_id in METADATA_POLICY_IDS:
        expected = f"metadata_policy:{decision_id}"
        if not any(row.get("decision_id") == expected for row in metadata_decisions):
            errors.append(f"MISSING_REQUIRED_METADATA_POLICY:{decision_id}")
    for row in data.get("uom_decisions", []):
        status = row.get("decision_status")
        if status not in DECISION_STATUSES:
            errors.append(f"INVALID_UOM_DECISION_STATUS:{row.get('profile_name')}:{row.get('raw_uom')}")
        if status == "approved" and (not row.get("decided_by") or not row.get("decided_at")):
            errors.append(f"APPROVED_UOM_REQUIRES_REVIEWER_AND_DATE:{row.get('profile_name')}:{row.get('raw_uom')}")
        if row.get("classification") == "ambiguous_unit" and status == "approved":
            if not row.get("selected_normalized_uom") or not row.get("rationale"):
                errors.append(f"AMBIGUOUS_UOM_REQUIRES_MAPPING_AND_RATIONALE:{row.get('profile_name')}:{row.get('raw_uom')}")
    if data.get("status") == "ready_to_apply":
        pending = [
            row.get("decision_id") or row.get("raw_uom")
            for row in all_profile_decisions + data.get("uom_decisions", [])
            if row.get("decision_status") == "pending"
        ]
        if pending:
            errors.append("READY_TO_APPLY_WITH_PENDING_DECISIONS")
    return errors


def extract_packet_answers(packet: Path) -> dict[str, str]:
    answers: dict[str, str] = {}
    current_id = ""
    for line in packet.read_text(encoding="utf-8").splitlines():
        if line.startswith("### "):
            current_id = line.removeprefix("### ").strip()
            continue
        if current_id and line.startswith("- Question:"):
            text = line.removeprefix("- Question:").strip()
            marker = "? "
            if marker in text:
                answer = text.split(marker, 1)[1].strip()
                if answer:
                    answers[current_id] = answer
    return answers


def normalize_packet_answer(answer: str, candidate_options: list[str]) -> str | None:
    cleaned = answer.strip()
    if not cleaned:
        return None
    if cleaned in candidate_options:
        return cleaned
    lowered = cleaned.lower()
    if lowered == "yes" and "nullable_allowed" in candidate_options:
        return "nullable_allowed"
    if lowered == "no" and "required_before_publish" in candidate_options:
        return "required_before_publish"
    for option in candidate_options:
        option_lower = option.lower()
        label = option.split("[", 1)[0].strip().lower()
        bracket = ""
        if "[" in option and "]" in option:
            bracket = option.split("[", 1)[1].split("]", 1)[0].strip().lower()
        if lowered == label or lowered == bracket:
            return option
        if lowered.startswith(label + " ") or (bracket and lowered.startswith(bracket + " ")):
            return option
    matches = [option for option in candidate_options if lowered in option.lower()]
    if len(matches) == 1:
        return matches[0]
    return None


def apply_packet_answers(data: dict[str, Any], packet: Path) -> dict[str, Any]:
    answers = extract_packet_answers(packet)
    applied: list[dict[str, str]] = []
    unmatched: list[dict[str, str]] = []
    decisions = data.get("profiles", []) + data.get("metadata_policy", {}).get("decisions", [])
    for row in decisions:
        decision_id = row.get("decision_id", "")
        if decision_id not in answers:
            continue
        selected = normalize_packet_answer(answers[decision_id], row.get("candidate_options", []))
        if not selected:
            unmatched.append({"decision_id": decision_id, "answer": answers[decision_id]})
            continue
        row["selected_option"] = selected
        row["decision_status"] = "pending"
        if selected != "unresolved" and not row.get("rationale"):
            row["rationale"] = "Captured from business_review_packet.md; pending reviewer approval."
        applied.append({"decision_id": decision_id, "selected_option": selected})
    data["updated_at"] = utc_now()
    return {"data": data, "applied": applied, "unmatched": unmatched, "answers_found": len(answers)}


def sync_packet(packet: Path, decisions: Path) -> dict[str, Any]:
    data = json.loads(decisions.read_text(encoding="utf-8"))
    result = apply_packet_answers(data, packet)
    decisions.write_text(json.dumps(result["data"], indent=2, sort_keys=True) + "\n", encoding="utf-8")
    validation_errors = validate_decisions(result["data"])
    return {
        "answers_found": result["answers_found"],
        "applied_count": len(result["applied"]),
        "applied": result["applied"],
        "unmatched": result["unmatched"],
        "valid": not validation_errors,
        "errors": validation_errors,
    }


def write_business_review_packet(
    path: Path,
    decisions_path: Path,
    profile_decisions: list[dict[str, Any]],
    uom_decisions: list[dict[str, Any]],
    detecto_presence: list[dict[str, Any]],
    metadata_rows: list[dict[str, Any]],
) -> None:
    lines = [
        "# Contract Pricing Business Review Packet",
        "",
        "This packet lists the business decisions required before any pilot profile can be approved or used for publishing active prices.",
        "",
        f"- Local decision file to edit: `{decisions_path}`",
        "- Validation command: `python -m pricing_ingestion.review.decisions validate --decisions data\\pricing\\business_decisions.local.json`",
        "- No prices, item identifiers, item descriptions, account numbers, contract numbers, or full source rows are included.",
        "",
        "## Required Decisions",
        "",
    ]
    for row in profile_decisions:
        lines.extend(
            [
                f"### {row['decision_id']}",
                f"- Category: {row['category']}",
                f"- Question: {row['question']}",
                f"- Candidate options: {', '.join(row['candidate_options'])}",
                f"- Evidence: {row['evidence_reference']}",
                "",
            ]
        )
    pending_uom = sum(1 for row in uom_decisions if row.get("decision_status") == "pending")
    lines.extend(
        [
            "## UOM Review",
            "",
            f"- UOM tokens requiring decisions: {pending_uom}",
            "- Direct lexical synonyms are proposed as candidates only; package expressions and ambiguous units remain pending.",
            "",
            "## Detecto Missing-Price Evidence",
            "",
        ]
    )
    if detecto_presence:
        for row in detecto_presence:
            lines.append(f"- {row['row_classification']}: {row['occurrence_count']} rows")
    else:
        lines.append("- No Detecto missing-price rows were found in the selected dry-run evidence.")
    lines.extend(["", "## Metadata Evidence", ""])
    for row in metadata_rows:
        lines.append(f"- {row['profile_name']}: {row['workbook_count']} workbook(s), publish requirements pending policy decisions.")
    lines.extend(
        [
            "",
            "## Unresolved Risks",
            "",
            "- No profile is approved.",
            "- Detecto missing-price rows remain blocking until a policy is approved.",
            "- Quantum Storage list-price semantics require explicit business confirmation.",
            "- Health Care Logistics price-UOM versus base-UOM semantics remain unresolved.",
            "- Database imports, item matching, package conversions, and each-price calculations are out of scope for this packet.",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def generate(
    review_root: Path,
    dry_run_root: Path,
    profiles_dir: Path,
    output: Path,
    business_review_output: Path = BUSINESS_REVIEW_OUTPUT,
    manifest: Path = Path("data/pricing/manifest.csv"),
    raw_root: Path = RAW_ROOT,
    uom_decision_path: Path = UOM_DECISION_PATH,
    force: bool = False,
) -> dict[str, Any]:
    runs = selected_runs(load_review_runs(dry_run_root, profiles_dir))
    business_review_output.mkdir(parents=True, exist_ok=True)

    detecto_presence, detecto_locations = build_detecto_price_presence_rows(runs, raw_root)
    price_candidates = build_price_basis_candidates(runs, raw_root)
    identifier_candidates = build_identifier_candidates(runs, raw_root)
    metadata_rows = build_metadata_completeness(runs, manifest)
    duplicate_rows = build_duplicate_review(runs)
    uom_decisions = build_uom_decisions(review_root)

    write_csv(business_review_output / "detecto_price_presence.csv", PRICE_PRESENCE_FIELDS, detecto_presence)
    write_csv(business_review_output / "detecto_manual_row_locations.csv", MANUAL_ROW_LOCATION_FIELDS, detecto_locations)
    write_csv(business_review_output / "price_basis_candidates.csv", PRICE_BASIS_FIELDS, price_candidates)
    write_csv(business_review_output / "identifier_candidates.csv", IDENTIFIER_FIELDS, identifier_candidates)
    write_csv(business_review_output / "uom_decision_template.csv", UOM_DECISION_FIELDS, uom_decisions)
    write_csv(business_review_output / "duplicate_review.csv", DUPLICATE_FIELDS, duplicate_rows)
    write_csv(business_review_output / "metadata_completeness.csv", METADATA_FIELDS, metadata_rows)

    existing = json.loads(output.read_text(encoding="utf-8")) if output.exists() else None
    decision_data = build_business_decisions(runs, review_root, price_candidates, identifier_candidates, existing)
    output_written = False
    if force or not output.exists():
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(json.dumps(decision_data, indent=2, sort_keys=True) + "\n", encoding="utf-8")
        output_written = True

    uom_existing_rows = read_csv(uom_decision_path) if uom_decision_path.exists() else []
    merged_uom = merge_uom_decisions(uom_decisions, uom_existing_rows)
    uom_local_written = False
    if force or not uom_decision_path.exists():
        write_csv(uom_decision_path, UOM_DECISION_FIELDS, merged_uom)
        uom_local_written = True

    packet_decisions = decision_data["profiles"] + decision_data["metadata_policy"]["decisions"]
    write_business_review_packet(business_review_output / "business_review_packet.md", output, packet_decisions, merged_uom, detecto_presence, metadata_rows)

    return {
        "runs_reviewed": len(runs),
        "decision_file": str(output),
        "decision_file_written": output_written,
        "uom_local_file": str(uom_decision_path),
        "uom_local_file_written": uom_local_written,
        "business_review_output": str(business_review_output),
        "profile_decisions": len(decision_data["profiles"]),
        "metadata_policy_decisions": len(decision_data["metadata_policy"]["decisions"]),
        "uom_tokens_pending": sum(1 for row in merged_uom if row.get("decision_status") == "pending"),
        "detecto_missing_price_classifications": dict(Counter(row["row_classification"] for row in detecto_presence for _ in range(int(row["occurrence_count"])))),
        "model_number_recommended": any(row.get("candidate_canonical_field") == "model_number" and row.get("profile_name", "").startswith("detecto") for row in identifier_candidates),
        "price_uom_base_uom_extension_recommended": True,
    }


def run_validate(decisions: Path) -> dict[str, Any]:
    data = json.loads(decisions.read_text(encoding="utf-8"))
    errors = validate_decisions(data)
    return {"valid": not errors, "errors": errors}


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Generate and validate contract-pricing business decisions.")
    subparsers = parser.add_subparsers(dest="command", required=True)
    generate_parser = subparsers.add_parser("generate")
    generate_parser.add_argument("--review-root", required=True, type=Path)
    generate_parser.add_argument("--dry-run-root", required=True, type=Path)
    generate_parser.add_argument("--profiles-dir", required=True, type=Path)
    generate_parser.add_argument("--output", required=True, type=Path)
    generate_parser.add_argument("--business-review-output", type=Path, default=BUSINESS_REVIEW_OUTPUT)
    generate_parser.add_argument("--manifest", type=Path, default=Path("data/pricing/manifest.csv"))
    generate_parser.add_argument("--raw-root", type=Path, default=RAW_ROOT)
    generate_parser.add_argument("--uom-decision-output", type=Path, default=UOM_DECISION_PATH)
    generate_parser.add_argument("--force", action="store_true")

    validate_parser = subparsers.add_parser("validate")
    validate_parser.add_argument("--decisions", required=True, type=Path)

    sync_parser = subparsers.add_parser("sync-packet")
    sync_parser.add_argument("--packet", required=True, type=Path)
    sync_parser.add_argument("--decisions", required=True, type=Path)
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_arg_parser()
    args = parser.parse_args(argv)
    if args.command == "generate":
        result = generate(args.review_root, args.dry_run_root, args.profiles_dir, args.output, args.business_review_output, args.manifest, args.raw_root, args.uom_decision_output, args.force)
        print(json.dumps(result, indent=2, sort_keys=True))
        return 0
    if args.command == "validate":
        result = run_validate(args.decisions)
        print(json.dumps(result, indent=2, sort_keys=True))
        return 0 if result["valid"] else 1
    if args.command == "sync-packet":
        result = sync_packet(args.packet, args.decisions)
        print(json.dumps(result, indent=2, sort_keys=True))
        return 0 if result["valid"] else 1
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
