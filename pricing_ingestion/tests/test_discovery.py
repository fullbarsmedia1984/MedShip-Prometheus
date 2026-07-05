from __future__ import annotations

import csv
import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from pricing_ingestion.discovery.discover import profile_workbook, run_discovery


class WorkbookDiscoveryTests(unittest.TestCase):
    def create_fixture_workbook(self, path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Contract Pricing"
        sheet["A1"] = "Synthetic Supplier Contract"
        sheet.merge_cells("A1:C1")
        sheet.append([])
        sheet.append(
            [
                "Vendor Part Number",
                "Manufacturer Part Number",
                "Description",
                "UOM",
                "Contract Price",
                "Effective Date",
                "Expiration Date",
            ]
        )
        sheet.append(["SKU-001", "MPN-001", "Synthetic item", "EA", 12.34, "2026-01-01", "2026-12-31"])
        sheet.append(["SKU-002", "MPN-002", "Another synthetic item", "CS", "=5+5", "2026-01-01", "2026-12-31"])
        sheet.row_dimensions[2].hidden = True
        sheet.column_dimensions["G"].hidden = True

        hidden_sheet = workbook.create_sheet("Internal Notes")
        hidden_sheet.sheet_state = "hidden"
        hidden_sheet["A1"] = "Synthetic hidden note"

        workbook.save(path)
        workbook.close()

    def test_profile_workbook_detects_structure_without_publishing_prices(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "synthetic_contract.xlsx"
            self.create_fixture_workbook(workbook_path)

            profile = profile_workbook(workbook_path, {"synthetic_contract.xlsx": {"supplier": "Synthetic"}})

            self.assertEqual(profile.file_name, "synthetic_contract.xlsx")
            self.assertEqual(profile.workbook_extension, ".xlsx")
            self.assertEqual(profile.manifest_metadata, {"supplier": "Synthetic"})
            self.assertIn("Internal Notes", profile.hidden_sheets)
            self.assertTrue(profile.candidate_pricing_sheets)

            pricing_sheet = next(sheet for sheet in profile.sheets if sheet.name == "Contract Pricing")
            self.assertEqual(pricing_sheet.hidden_row_count, 1)
            self.assertEqual(pricing_sheet.hidden_column_count, 1)
            self.assertEqual(pricing_sheet.formula_count, 1)
            self.assertIn("contains_merged_cells", pricing_sheet.risk_flags)
            self.assertTrue(pricing_sheet.likely_distributor_sku_columns)
            self.assertTrue(pricing_sheet.likely_manufacturer_part_number_columns)
            self.assertTrue(pricing_sheet.likely_description_columns)
            self.assertTrue(pricing_sheet.likely_uom_columns)
            self.assertTrue(pricing_sheet.likely_price_columns)
            self.assertTrue(pricing_sheet.likely_effective_date_columns)
            self.assertTrue(pricing_sheet.likely_expiration_date_columns)

            sample_values = pricing_sheet.sample_rows[0]["values"]
            self.assertEqual(sample_values["E"], "<redacted_price_value>")
            self.assertNotIn("12.34", json.dumps(profile, default=str))

    def test_run_discovery_writes_profiles_and_summary_reports(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "raw"
            output_dir = root / "out"
            input_dir.mkdir()

            workbook_path = input_dir / "synthetic_contract.xlsx"
            self.create_fixture_workbook(workbook_path)

            manifest_path = root / "manifest.csv"
            with manifest_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(handle, fieldnames=["file_name", "supplier"])
                writer.writeheader()
                writer.writerow({"file_name": "synthetic_contract.xlsx", "supplier": "Synthetic"})

            profiles = run_discovery(input_dir, manifest_path, output_dir)

            self.assertEqual(len(profiles), 1)
            self.assertTrue((output_dir / "workbook_inventory.csv").exists())
            self.assertTrue((output_dir / "sheet_inventory.csv").exists())
            self.assertTrue((output_dir / "candidate_headers.csv").exists())
            self.assertTrue((output_dir / "risk_report.md").exists())
            self.assertTrue((output_dir / "discovery_summary.md").exists())

            profile_files = list((output_dir / "profiles").glob("*.pricing_discovery.json"))
            self.assertEqual(len(profile_files), 1)
            profile_json = json.loads(profile_files[0].read_text(encoding="utf-8"))
            self.assertEqual(profile_json["file_name"], "synthetic_contract.xlsx")
            self.assertEqual(profile_json["manifest_metadata"]["supplier"], "Synthetic")

            with (output_dir / "candidate_headers.csv").open(encoding="utf-8") as handle:
                rows = list(csv.DictReader(handle))
            self.assertEqual(rows[0]["sheet_name"], "Contract Pricing")

    def test_xls_is_reported_as_adapter_gap_without_parsing(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "legacy_contract.xls"
            path.write_bytes(b"synthetic legacy workbook placeholder")

            profile = profile_workbook(path, {})

            self.assertEqual(profile.workbook_extension, ".xls")
            self.assertIn("xls_not_supported_without_xlrd_adapter", profile.risk_flags)
            self.assertEqual(profile.sheets, [])


if __name__ == "__main__":
    unittest.main()

