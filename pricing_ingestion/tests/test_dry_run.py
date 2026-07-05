from __future__ import annotations

import csv
import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from pricing_ingestion.dry_run.runner import run_dry_run
from pricing_ingestion.tests.test_profiles import valid_profile


class DryRunTests(unittest.TestCase):
    def write_profile(self, path: Path, profile: dict | None = None) -> None:
        path.write_text(json.dumps(profile or valid_profile()), encoding="utf-8")

    def create_workbook(self, path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Pricing"
        sheet.append(["SKU", "Description", "UOM", "Price"])
        sheet.append(["001", "First item", "EA", "$10.00"])
        sheet.append(["SKU", "Description", "UOM", "Price"])
        sheet.append(["002", "Second item", "MYSTERY", "$12.00"])
        sheet.append(["003", "Third item", "BX", "=5+5"])
        sheet.append(["HIDDEN", "Hidden item", "EA", "$9.00"])
        sheet.append(["004", "Fourth item", "EA", "$10.00"])
        sheet.append(["004", "Fourth item", "EA", "$10.00"])
        sheet.append(["004", "Fourth item", "EA", "$11.00"])
        sheet.append(["", "Section Heading", "", ""])
        sheet.append(["005", "", "EA", "$13.00"])
        sheet.append(["", "", "", ""])
        sheet.append(["", "", "", ""])
        sheet.row_dimensions[6].hidden = True
        sheet.column_dimensions["D"].hidden = True
        sheet.append(["Grand Total", "", "", "$99.00"])
        sheet.merge_cells("A12:B12")
        sheet["A12"] = "Merged note"
        workbook.save(path)
        workbook.close()

    def test_dry_run_extracts_rows_with_lineage_and_safe_markdown(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            workbook_path = root / "Synthetic Pricing.xlsx"
            profile_path = root / "profile.json"
            output_dir = root / "out"
            self.create_workbook(workbook_path)
            self.write_profile(profile_path)

            summary = run_dry_run(workbook_path, profile_path, None, output_dir)

            self.assertGreater(summary["rows_scanned"], 0)
            self.assertTrue((output_dir / "proposed_rows.csv").exists())
            self.assertTrue((output_dir / "exceptions.csv").exists())
            self.assertTrue((output_dir / "dry_run_summary.md").exists())
            self.assertTrue((output_dir / "mapping_review.md").exists())
            self.assertTrue((output_dir / "excluded_rows.csv").exists())

            with (output_dir / "proposed_rows.csv").open(encoding="utf-8") as handle:
                rows = list(csv.DictReader(handle))
            self.assertTrue(rows[0]["source_file"])
            self.assertEqual(rows[0]["source_sheet_name"], "Pricing")
            self.assertEqual(rows[0]["source_row_number"], "2")
            self.assertIn("distributor_sku", rows[0]["source_column_map"])
            self.assertIn("distributor_sku", rows[0]["source_cell_map"])

            markdown = (output_dir / "dry_run_summary.md").read_text(encoding="utf-8")
            self.assertNotIn("$10.00", markdown)
            self.assertNotIn("First item", markdown)

            with (output_dir / "exceptions.csv").open(encoding="utf-8") as handle:
                exception_codes = {row["exception_code"] for row in csv.DictReader(handle)}
            self.assertIn("UNKNOWN_UOM", exception_codes)
            self.assertIn("FORMULA_WITHOUT_CACHED_VALUE", exception_codes)
            self.assertIn("DUPLICATE_IDENTICAL_ROW", exception_codes)
            self.assertIn("DUPLICATE_CONFLICTING_PRICE", exception_codes)
            self.assertIn("MISSING_DESCRIPTION", exception_codes)

            with (output_dir / "excluded_rows.csv").open(encoding="utf-8") as handle:
                excluded_reasons = {row["reason"] for row in csv.DictReader(handle)}
            self.assertIn("SECTION_OR_NON_ITEM_ROW_SKIPPED", excluded_reasons)

    def test_candidate_header_row_range(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            workbook_path = root / "Synthetic Pricing.xlsx"
            profile_path = root / "profile.json"
            output_dir = root / "out"
            self.create_workbook(workbook_path)
            profile = valid_profile()
            profile["header_rules"]["exact_header_row"] = None
            profile["header_rules"]["candidate_row_range"] = {"start": 1, "end": 3}
            profile["header_rules"]["minimum_header_match_score"] = 4
            profile["data_region_rules"]["explicit_first_data_row"] = None
            profile["data_region_rules"]["first_row_after_detected_header"] = True
            self.write_profile(profile_path, profile)

            summary = run_dry_run(workbook_path, profile_path, None, output_dir)
            self.assertGreater(summary["proposed_rows"], 0)

    def test_fallback_source_header_is_used_when_primary_blank(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            workbook_path = root / "Synthetic Pricing.xlsx"
            profile_path = root / "profile.json"
            output_dir = root / "out"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Pricing"
            sheet.append(["SKU", "Description", "Alt Description", "UOM", "Price"])
            sheet.append(["001", "", "Fallback description", "EA", "$10.00"])
            workbook.save(workbook_path)
            workbook.close()

            profile = valid_profile()
            profile["column_mappings"][1]["source_header"] = "Description"
            profile["column_mappings"][1]["source_column_letter"] = "B"
            profile["column_mappings"][1]["source_column_index"] = 2
            profile["column_mappings"][1]["fallback_source_headers"] = ["Alt Description"]
            profile["column_mappings"][2]["source_column_letter"] = "D"
            profile["column_mappings"][2]["source_column_index"] = 4
            profile["column_mappings"][3]["source_column_letter"] = "E"
            profile["column_mappings"][3]["source_column_index"] = 5
            self.write_profile(profile_path, profile)

            summary = run_dry_run(workbook_path, profile_path, None, output_dir)
            self.assertEqual(summary["blocking_exception_rows"], 0)

    def test_model_number_price_uom_and_metadata_source_tracking(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            workbook_path = root / "Synthetic Pricing.xlsx"
            profile_path = root / "profile.json"
            manifest_path = root / "manifest.csv"
            output_dir = root / "out"

            self.create_workbook(workbook_path)
            profile = valid_profile()
            profile["column_mappings"][0]["canonical_field"] = "model_number"
            price_uom = dict(profile["column_mappings"][2])
            price_uom["canonical_field"] = "raw_price_uom"
            price_uom["required"] = False
            profile["column_mappings"].append(price_uom)
            profile["default_values"]["contract_number"] = {"source": "manifest", "value": None, "manifest_field": "contract_number", "required": True}
            profile["default_values"]["effective_date"] = {"source": "manifest", "value": None, "manifest_field": "effective_date", "required": True}
            profile["validation_rules"]["required_metadata_fields"] = ["contract_number", "effective_date"]
            self.write_profile(profile_path, profile)
            with manifest_path.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=["file_name", "contract_number", "effective_date"])
                writer.writeheader()
                writer.writerow({"file_name": workbook_path.name, "contract_number": "FICTIONAL-CONTRACT", "effective_date": "2026-01-01"})

            summary = run_dry_run(workbook_path, profile_path, manifest_path, output_dir)
            self.assertGreater(summary["proposed_rows"], 0)
            with (output_dir / "proposed_rows.csv").open(encoding="utf-8") as handle:
                rows = list(csv.DictReader(handle))
            self.assertIn("model_number", rows[0])
            self.assertIn("raw_price_uom", rows[0])
            self.assertIn("normalized_price_uom", rows[0])
            self.assertIn("metadata_source_map", rows[0])
            self.assertIn("manifest", rows[0]["metadata_source_map"])

    def test_required_metadata_missing_blocks_rows(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            workbook_path = root / "Synthetic Pricing.xlsx"
            profile_path = root / "profile.json"
            output_dir = root / "out"

            self.create_workbook(workbook_path)
            profile = valid_profile()
            profile["default_values"]["contract_number"] = {"source": "manifest", "value": None, "manifest_field": "contract_number", "required": True}
            profile["validation_rules"]["required_metadata_fields"] = ["contract_number"]
            self.write_profile(profile_path, profile)

            summary = run_dry_run(workbook_path, profile_path, None, output_dir)
            self.assertGreater(summary["exception_counts"].get("MISSING_REQUIRED_CONTRACT_METADATA", 0), 0)


if __name__ == "__main__":
    unittest.main()
